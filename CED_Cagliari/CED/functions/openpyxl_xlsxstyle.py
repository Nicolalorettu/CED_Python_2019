from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.utils.dataframe import dataframe_to_rows

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None, union=None, grid=False, interborder=False, width=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    :param union: An openpyxl Merge object
    :param alignment: An openpyxl Positional object
    :param interborder: An openpyxl internal Border
    :param columnwidth: Excel column letter to apply new width dimension
    :param width: Column width dimension
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    if cell_range.find(":") == -1:
        cell_range = cell_range + ":" + cell_range

    first_cell = ws[cell_range.split(":")[0]]


    if union:
        ws.merge_cells(cell_range)

    if alignment:
        for row in ws[cell_range]:
            for cell in row:
                cell.alignment = alignment

    rows = ws[cell_range]


    if font:
        for cell in rows[0]:
            cell.font = font

    if interborder is False:
        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom
    else:
        for cell in rows[0]:
            cell.border = cell.border + top
            cell.border = cell.border + left
        for cell in rows[-1]:
            cell.border = cell.border + bottom
            cell.border = cell.border + right

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

    if grid is not True:
        ws.sheet_view.showGridLines = False
    else:
        ws.sheet_view.showGridLines = True

    if width is not None:
        for cell in rows[0]:
            init = str(cell).find("'.")+2
            end = init+1
            ws.column_dimensions[str(cell)[init:end]].width = width



def df_to_sheet(wb, df, sheetname, index=True):
    ws = wb.create_sheet(sheetname)
    for r in dataframe_to_rows(df, index=index, header=True):
        ws.append(r)

    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
    return wb
