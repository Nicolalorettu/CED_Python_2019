from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Color
from openpyxl import load_workbook, Workbook
import openpyxl
import pandas as pd
try:
    from openpyxl_xlsxstyle import style_range, df_to_sheet
except ImportError:
    from .openpyxl_xlsxstyle import style_range, df_to_sheet

thin = Side(border_style="thin", color="000000")
medium = Side(border_style="medium", color="000000")
nobord = Side(border_style=None)
no_fill = PatternFill(fill_type=None)

borderempty = Border(top=nobord, left=nobord, right=nobord, bottom=nobord)
borderthin = Border(top=thin, left=thin, right=thin, bottom=thin)
borderthinlat = Border(top=nobord, left=thin, right=thin, bottom=nobord)
borderthinbot = Border(top=nobord, left=thin, right=thin, bottom=thin)
bordermedium = Border(top=medium, left=medium, right=medium, bottom=medium)

aligncenter = Alignment(horizontal="center", vertical="center", wrap_text=True)
alignvertcenter = Alignment(horizontal="left", vertical="center", wrap_text=True)
fillred = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
fillgray = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')
filllighgray = PatternFill(start_color='f0f0f0', end_color='f0f0f0', fill_type='solid')
fillpink = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')
fillgreen = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
fillazure = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
fillblue= PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')

fontboldwhite = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FFFFFF')
fontboldblack = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
fontblack = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='000000')
fontwhite = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FFFFFF')


def Atpay_easy_xls_template(htmlth, table, smsmonth, smsweek, datef, datel):
    wb = Workbook()
    ws = wb.active
    cellsrangemonth = ["B5:L5", "B6:L6", "B7:L7"]
    bordermonth = [borderempty, borderthin, borderthin]
    fillcellmonth = [fillred, fillgray, no_fill]
    fonttablemonth = [fontboldblack, fontblack, fontblack]
    uniontablemonth = [True, False, False]
    # Upper Title
    enlargedcol = ["B", "I", "J", "K"]
    for i in enlargedcol:
        ws.column_dimensions[i].width = 12.78
    ws.column_dimensions["A"].width = 2.78

    img = openpyxl.drawing.image.Image(r"C:\Apache\htdocs\CED_Cagliari\CED_Cagliari\static\app\img\excel_logo.png")
    img.anchor = "B2"
    ws.add_image(img)
    ws['D2'] = "Estrazione Dati dal " + datef + " al " + datel
    title = "D2:H3"
    style_range(ws, title, border=borderempty, fill=no_fill, font=fontblack, alignment=aligncenter, union=True, grid=False, interborder=False, width=None)
    # Dati del Mese
    ws['B5'].value = "MENSILE"
    Intestazione = ws[cellsrangemonth[1]]
    i = 0
    for cell in Intestazione[0]:
        cell.value = htmlth[i]
        i += 1
    tuplemcount = (len(smsmonth[table[2]]))
    tuplewcount = (len(smsweek[table[2]]))
    for k in range(0, 3):
        style_range(ws, cellsrangemonth[k], border=bordermonth[k], fill=fillcellmonth[k], font=fonttablemonth[k], alignment=aligncenter, union=uniontablemonth[k], grid=False, interborder=True)
        if k == 2:
            fcolumn = cellsrangemonth[k][:1]
            lcolumn = cellsrangemonth[k][3]
            for m in range(0, tuplemcount):
                frows = 7 + m
                multimonth = fcolumn + str(frows) + ":" + lcolumn + str(frows)
                style_range(ws, multimonth, border=bordermonth[2], fill=fillcellmonth[2], font=fonttablemonth[2], alignment=aligncenter, union=uniontablemonth[2], grid=False, interborder=True)
                valuesmonth = list(smsmonth[table[2]][m])
                j = 0
                rows = ws[multimonth]
                for cell in rows[j]:
                    cell.value = smsmonth[table[2]][m][valuesmonth[j]]
                    j += 1
    # empty row
    lastrow = str(ws[multimonth][-1])[16:17]
    fwcolumn = 'B'
    lwcolumn = 'L'
    emptyrow = fwcolumn + str(int(lastrow)+1) + ":" + lwcolumn + str(int(lastrow)+1)
    style_range(ws, emptyrow, border=bordermonth[0], fill=None, font=None, alignment=None, union=uniontablemonth[2], grid=False, interborder=False)

    # Dati delle Weeks
    for n in range(2, 3):
        intestazioneweek = fwcolumn + str(int(lastrow)+n)
        ws[intestazioneweek].value = "SETTIMANALE"
        rowintestazione = intestazioneweek + ":" + lwcolumn + str(int(lastrow)+n)
        style_range(ws, rowintestazione, border=bordermonth[0], fill=fillcellmonth[0], font=fonttablemonth[0], alignment=aligncenter, union=uniontablemonth[0], grid=False, interborder=True)
        n += 1
        if n == 3:
            rowintestazione = fwcolumn + str(int(lastrow)+n) + ":" + lwcolumn + str(int(lastrow)+n)
            o = 0
            rows = ws[rowintestazione]
            for cell in rows[o]:
                cell.value = htmlth[o]
                style_range(ws, rowintestazione, border=bordermonth[1], fill=fillcellmonth[1], font=fonttablemonth[1], alignment=aligncenter, union=uniontablemonth[1], grid=False, interborder=True)
                o += 1
            n += 1
            for p in range(0, tuplewcount):
                frows = (int(lastrow)+n+p)
                multiweek = fcolumn + str(frows) + ":" + lcolumn + str(frows)
                style_range(ws, multiweek, border=bordermonth[2], fill=fillcellmonth[2], font=fonttablemonth[2], alignment=aligncenter, union=uniontablemonth[2], grid=False, interborder=True)
                valuesweek = list(smsweek[table[2]][p])
                periodweeks = "Week " + str(smsweek[table[2]][p][valuesweek[0]])
                j = 0
                rows = ws[multiweek]
                for cell in rows[j]:
                    cell.value = (smsweek[table[2]][p][valuesweek[j]])
                    j += 1
                refillperiodweek = fwcolumn + str(int(lastrow)+n+p)
                column = ws[refillperiodweek]
                column.value = periodweeks
    return wb


def atpay_easy_paf(data, days, supervisor, discount, site):
    path = r"C:\Apache\htdocs\CED_Cagliari\CED\xls_templates\atpay_paf.xlsx"
    wb = load_workbook(path)
    tkts = list(data)
    ws = wb.active
    typedistance = 5 + len(data["tkts"])
    headerdistance = 10
    cellsgeneral = ["B2", "AF2", "A10", "B3", "D4", "D5", "G2", "G3", "AH3", "AH2"]
    monthlist = list(data["total"])
    cellstkts = list()
    for i in range(0, len(cellsgeneral)):
        if i == 0:
            ws[cellsgeneral[i]].value = supervisor
        elif i == 1:
            ws[cellsgeneral[i]].value = discount
        elif i == 2:
            ws[cellsgeneral[i]].value = site
        else:
            ws[cellsgeneral[i]].value = data["total"][monthlist[i-2]]

    for i in range(0, len(data["tkts"])):
        ws["I"+str(headerdistance+i+(typedistance*0))].value = data["tkts"][i]["month"]["preventivato_pezzi"]
        ws["J"+str(headerdistance+i+(typedistance*0))].value = data["tkts"][i]["month"]["preventivato_euro"] - data["tkts"][i]["month"]["quota_accantonamento"]
        ws["I"+str(headerdistance+i+(typedistance*1))].value = data["tkts"][i]["month"]["consuntivato_pezzi"]
        ws["J"+str(headerdistance+i+(typedistance*1))].value = data["tkts"][i]["month"]["consuntivato_euro"]
        ws["I"+str(headerdistance+i+(typedistance*2))].value = data["tkts"][i]["month"]["preventivato_pezzi"] - data["tkts"][i]["month"]["consuntivato_pezzi"]
        ws["J"+str(headerdistance+i+(typedistance*2))].value = data["tkts"][i]["month"]["preventivato_euro"] - data["tkts"][i]["month"]["consuntivato_euro"] - data["tkts"][i]["month"]["quota_accantonamento"]
        for k in range(0, 3):
            ws["B"+str(headerdistance+i+(typedistance*k))].value = data["tkts"][i]["service"]
            cellstkts.append("B"+str(headerdistance+i+(typedistance*k))+":C"+str(headerdistance+i+(typedistance*k)))
            ws["F"+str(headerdistance+i+(typedistance*k))].value = data["tkts"][i]["tktvalue"]
        for col in range(11, 42-(31-int(days))):
            ws.cell(column=col, row=headerdistance+(typedistance*0)-1, value=data["tkts"][i]["daily"][col-11]["day"])
            ws.cell(column=col, row=headerdistance+i+(typedistance*0), value=data["tkts"][i]["daily"][col-11]["ok"])
            if data["total"]["last_day"] >= (col-10):
                ws.cell(column=col, row=headerdistance+i+(typedistance*1), value=data["tkts"][i]["daily"][col-11]["ok"])
                ws.cell(column=col, row=headerdistance+i+(typedistance*2), value=0)
            elif data["total"]["last_day"] < (col-10):
                ws.cell(column=col, row=headerdistance+i+(typedistance*2), value=data["tkts"][i]["daily"][col-11]["ok"])
                ws.cell(column=col, row=headerdistance+i+(typedistance*1), value=0)

    month = dict()
    month["28"] = "AL"
    month["30"] = "AN"
    month["31"] = "AO"

    cellsrange = ["A1:E1", "F1:G1", "AC1:AJ1", "AC2:AE2", "AF2:AG2", "AH2:AJ2", "AC3:AE3", "AF3:AG3", "AH3:AJ3",                                    # indice 8
                  "D2:E2", "D3:E3", "D4:E4", "D5:E5"]                                                                                               # indice 12

    for i in range(0, 3):
        cellstkts.append("D"+str(headerdistance+(typedistance*i)-3)+":F"+str(headerdistance+(typedistance*i)-3))
        cellstkts.append("G"+str(headerdistance+(typedistance*i)-3)+":H"+str(headerdistance+(typedistance*i)-3))
        cellstkts.append("I"+str(headerdistance+(typedistance*i)-3)+":J"+str(headerdistance+(typedistance*i)-3))

        cellstkts.append("K"+str(headerdistance+(typedistance*i)-3)+":"+month[days]+str(headerdistance+(typedistance*i)-3))

        cellstkts.append("B"+str(headerdistance+(typedistance*i)-2)+":C"+str(headerdistance+(typedistance*i)-1))
        cellstkts.append("D"+str(headerdistance+(typedistance*i)-2)+":D"+str(headerdistance+(typedistance*i)-1))
        cellstkts.append("E"+str(headerdistance+(typedistance*i)-2)+":E"+str(headerdistance+(typedistance*i)-1))
        cellstkts.append("F"+str(headerdistance+(typedistance*i)-2)+":F"+str(headerdistance+(typedistance*i)-1))
        cellstkts.append("G"+str(headerdistance+(typedistance*i)-2)+":G"+str(headerdistance+(typedistance*i)-1))
        cellstkts.append("H"+str(headerdistance+(typedistance*i)-2)+":H"+str(headerdistance+(typedistance*i)-1))
        cellstkts.append("I"+str(headerdistance+(typedistance*i)-2)+":I"+str(headerdistance+(typedistance*i)-1))
        cellstkts.append("J"+str(headerdistance+(typedistance*i)-2)+":J"+str(headerdistance+(typedistance*i)-1))

        cellstkts.append("A"+str(headerdistance+(typedistance*i))+":A"+str(headerdistance+len(data["tkts"])-1+(typedistance*i)))

    for i in range(len(cellstkts)):
        cellsrange.append(cellstkts[i])
    for i in range(len(cellsrange)):
        if i > 2 and i <= 8:
            border = Border(top=thin, left=medium, right=medium, bottom=thin)
        elif i > 8 and i <= 12:
            border = Border()
        elif i > 12 and i <= 12+(len(data["tkts"])*3):
            border = Border(top=thin, left=medium, right=medium, bottom=thin)
        else:
            border = Border(top=medium, left=medium, right=medium, bottom=medium)

        style_range(ws, cellsrange[i], border=border, fill=None, font=None, alignment=aligncenter, union=True, grid=False, interborder=False, width=None)

    if int(days) < 31:
        ws.delete_cols((38+31-int(days)), 41)
    return wb


def c87_ivr_xls_template(table, ivrmonth, ivrweek, htmlth, datefrom, dateto, df1, df2, tgt):
    wb = Workbook()
    ws = wb.active

    reductedcol = ["A", "D", "L", "S"]
    for i in reductedcol:
        ws.column_dimensions[i].width = 2.78

    try:
        img = openpyxl.drawing.image.Image(r"C:\Apache\htdocs\CED_Cagliari\CED_Cagliari\static\app\img\excel_logo.png")
        img.anchor = "B2"
        ws.add_image(img)
    except FileNotFoundError:
        None
    ws["E2"] = "Estrazione dal " + datefrom + " al " + dateto
    ws["B5"] = "MENSILE"
    ws["E5"] = "187 HOME"
    ws["M5"] = "191 OFFICE"

    style_range(ws, "E2:H3", border=borderempty, fill=no_fill, font=fontboldblack, alignment=alignvertcenter, union=True, grid=False, interborder=False, width=None)

    headercol = ["B5:C5", "E5:K5", "M5:R5"]
    for cells in headercol:
        style_range(ws, cells, border=borderthin, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=True, grid=False, interborder=False, width=None)

    headercol = ["B6:C6", "E6:K6", "M6:R6"]
    for cells in headercol:
        style_range(ws, cells, border=borderthin, fill=fillgray, font=fontboldblack, alignment=aligncenter, union=False, grid=False, interborder=True, width=10.78)

    headerdistance = 6
    x = 0
    y = 0
    for k in range(0, 17):
        if k < 2:
            ws.cell(column=k+2, row=headerdistance, value=htmlth["int"][k])
        elif k == 2 or k == 10:
            None
        elif k < 10 and k > 2:
            ws.cell(column=k+2, row=headerdistance, value=htmlth["187new"][x])
            x += 1
        elif k > 10:
            ws.cell(column=k+2, row=headerdistance, value=htmlth["191new"][y])
            y += 1

    columlist187 = list(ivrmonth[table[0]][0][0])
    columlist191 = list(ivrmonth[table[1]][0][0])
    for i in range(0, len(ivrmonth[table[0]][0])):
        x = 2
        y = 2
        col = ["B" + str(headerdistance+i+1) + ":C" + str(headerdistance+i+1),"E" + str(headerdistance+i+1) + ":K" + str(headerdistance+i+1), "M" + str(headerdistance+i+1) + ":R" + str(headerdistance+i+1)]
        if (i % 2) == 0:
            if i == (len(ivrmonth[table[0]][0]) - 1):
                for cells in col:
                    style_range(ws, cells, border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for cells in col:
                style_range(ws, cells, border=borderthinlat, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
        else:
            if i == (len(ivrmonth[table[0]][0]) - 1):
                for cells in col:
                    style_range(ws, cells, border=borderthinbot, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for cells in col:
                style_range(ws, cells, border=borderthinlat, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
        for k in range(0, 17):
            if k < 2:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrmonth[table[0]][0][i][columlist187[k]])
            elif k == 2 or k == 10:
                None
            elif k > 2 and k < 6:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrmonth[table[0]][0][i][columlist187[x]])
                x += 1
            elif k >= 6 and k < 9:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrmonth[table[0]][0][i][columlist187[x]])
                if tgt["home"][k-6]["tier3"] == 0:
                    None
                elif ivrmonth[table[0]][0][i][columlist187[x]] > tgt["home"][k-6]["tier3"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillblue, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrmonth[table[0]][0][i][columlist187[x]] > tgt["home"][k-6]["tier2"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillazure, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrmonth[table[0]][0][i][columlist187[x]] > tgt["home"][k-6]["tier1"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrmonth[table[0]][0][i][columlist187[x]] > tgt["home"][k-6]["tgt"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                else:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                x += 1
            elif k == 9:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrmonth[table[0]][0][i][columlist187[18]])
                if tgt["home"][k-6]["tier3"] == 0:
                    None
                elif ivrmonth[table[0]][0][i][columlist187[18]] > tgt["home"][k-6]["tier3"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillblue, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrmonth[table[0]][0][i][columlist187[18]] > tgt["home"][k-6]["tier2"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillazure, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrmonth[table[0]][0][i][columlist187[18]] > tgt["home"][k-6]["tier1"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrmonth[table[0]][0][i][columlist187[18]] > tgt["home"][k-6]["tgt"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                else:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            elif k > 10 and k < 14:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrmonth[table[1]][0][i][columlist191[y]])
                y += 1
            elif k >= 14:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrmonth[table[1]][0][i][columlist191[y]])
                if tgt["office"][k-14]["tier3"] == 0:
                    None
                elif ivrmonth[table[1]][0][i][columlist191[y]] > tgt["office"][k-14]["tier3"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillblue, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrmonth[table[1]][0][i][columlist191[y]] > tgt["office"][k-14]["tier2"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillazure, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrmonth[table[1]][0][i][columlist191[y]] > tgt["office"][k-14]["tier1"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrmonth[table[1]][0][i][columlist191[y]] > tgt["office"][k-14]["tgt"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                else:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:16], border=borderempty, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                y += 1
    headerdistance += i+3

    ws["B" + str(headerdistance)] = "SETTIMANALE"
    ws["E" + str(headerdistance)] = "187 HOME"
    ws["M" + str(headerdistance)] = "191 OFFICE"

    headercol = ["B" + str(headerdistance) +":C" + str(headerdistance), "E" + str(headerdistance) +":K" + str(headerdistance), "M" + str(headerdistance) +":R" + str(headerdistance)]
    for cells in headercol:
        style_range(ws, cells, border=borderthin, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=True, grid=False, interborder=False, width=None)

    headerdistance += 1
    headercol = ["B" + str(headerdistance) +":C" + str(headerdistance), "E" + str(headerdistance) +":K" + str(headerdistance), "M" + str(headerdistance) +":R" + str(headerdistance)]

    for cells in headercol:
        style_range(ws, cells, border=borderthin, fill=fillgray, font=fontboldblack, alignment=aligncenter, union=False, grid=False, interborder=True, width=10.78)
    x = 0
    y = 0
    for k in range(0, 17):
        if k < 2:
            ws.cell(column=k+2, row=headerdistance, value=htmlth["int"][k])
        elif k == 2 or k == 10:
            None
        elif k < 10 and k > 2:
            ws.cell(column=k+2, row=headerdistance, value=htmlth["187new"][x])
            x += 1
        elif k > 10:
            ws.cell(column=k+2, row=headerdistance, value=htmlth["191new"][y])
            y += 1

    columlist187 = list(ivrweek[table[0]][0][1])
    columlist191 = list(ivrweek[table[1]][0][1])
    for i in range(0, len(ivrweek[table[0]][0])):
        x = 2
        y = 2
        col = ["B" + str(headerdistance+i+1) + ":C" + str(headerdistance+i+1),"E" + str(headerdistance+i+1) + ":K" + str(headerdistance+i+1), "M" + str(headerdistance+i+1) + ":R" + str(headerdistance+i+1)]
        if (i % 2) == 0:
            if i == (len(ivrweek[table[0]][0]) - 1):
                for cells in col:
                    style_range(ws, cells, border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for cells in col:
                style_range(ws, cells, border=borderthinlat, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
        else:
            if i == (len(ivrweek[table[0]][0]) - 1):
                for cells in col:
                    style_range(ws, cells, border=borderthinbot, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for cells in col:
                style_range(ws, cells, border=borderthinlat, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
        for k in range(0, 17):
            if k < 2:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrweek[table[0]][0][i][columlist187[k]])
            elif k == 2 or k == 10:
                None
            elif k > 2 and k < 6:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrweek[table[0]][0][i][columlist187[x]])
                x += 1
            elif k >= 6 and k < 9:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrweek[table[0]][0][i][columlist187[x]])
                if tgt["home"][k-6]["tier3"] == 0:
                    None
                elif ivrweek[table[0]][0][i][columlist187[x]] > tgt["home"][k-6]["tier3"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillblue, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrweek[table[0]][0][i][columlist187[x]] > tgt["home"][k-6]["tier2"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillazure, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrweek[table[0]][0][i][columlist187[x]] > tgt["home"][k-6]["tier1"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrweek[table[0]][0][i][columlist187[x]] > tgt["home"][k-6]["tgt"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                else:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                x += 1
            elif k == 9:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrweek[table[0]][0][i][columlist187[18]])
                if tgt["home"][k-6]["tier3"] == 0:
                    None
                elif ivrweek[table[0]][0][i][columlist187[18]] > tgt["home"][k-6]["tier3"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillblue, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrweek[table[0]][0][i][columlist187[18]] > tgt["home"][k-6]["tier2"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillazure, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrweek[table[0]][0][i][columlist187[18]] > tgt["home"][k-6]["tier1"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrweek[table[0]][0][i][columlist187[18]] > tgt["home"][k-6]["tgt"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                else:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            elif k > 10 and k < 14:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrweek[table[1]][0][i][columlist191[y]])
                y += 1
            elif k >= 14:
                ws.cell(column=k+2, row=headerdistance+i+1, value=ivrweek[table[1]][0][i][columlist191[y]])
                if tgt["office"][k-14]["tier3"] == 0:
                    None
                elif ivrweek[table[1]][0][i][columlist191[y]] > tgt["office"][k-14]["tier3"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillblue, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrweek[table[1]][0][i][columlist191[y]] > tgt["office"][k-14]["tier2"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillazure, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrweek[table[1]][0][i][columlist191[y]] > tgt["office"][k-14]["tier1"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                elif ivrweek[table[1]][0][i][columlist191[y]] > tgt["office"][k-14]["tgt"]:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                else:
                    style_range(ws, str(ws.cell(column=k+2, row=headerdistance+i+1))[14:17], border=borderempty, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                y += 1

    ws.title = "Riepilogo"

    wb = df_to_sheet(wb, df1, table[0][4:])
    wb = df_to_sheet(wb, df2, table[1][4:])
    return wb


def c87_opera_xls_template(form, opera_kpi, tgt, df, indexdf):
    wb = Workbook()
    ws = wb.active

    headerdistance = 6
    reductedcol = ["A", "F"]

    for i in reductedcol:
        ws.column_dimensions[i].width = 2.78

    try:
        img = openpyxl.drawing.image.Image(r"C:\Apache\htdocs\CED_Cagliari\CED_Cagliari\static\app\img\excel_logo.png")
        img.anchor = "B2"
        ws.add_image(img)
    except FileNotFoundError:
        None

    ws["E2"] = "Estrazione di " + form["month"] + " " + form["year"]
    ws["B5"] = "Report Sintesi Volumi FE"
    ws["J5"] = "Report Sintesi Volumi BO"
    style_range(ws, "E2:H3", border=borderempty, fill=no_fill, font=fontboldblack, alignment=alignvertcenter, union=True, grid=False, interborder=False, width=None)

    headercol = ["B5:H5", "J5:P5"]
    for cells in headercol:
        style_range(ws, cells, border=borderthin, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=True, grid=False, interborder=False, width=None)

    headercol = ["B6", "C6:H6", "J6", "K6:P6"]
    for cells in headercol:
        if cells == "B6" or cells == "J6":
            width = 21.78
        else:
            width = 8.78
        style_range(ws, cells, border=borderthin, fill=fillgray, font=fontboldblack, alignment=aligncenter, union=False, grid=False, interborder=True, width=width)

    header = ["INBOUND HOME", "TARGET <", "TIER1", "TIER2", "TIER3", "ENNOVA","Delta vs KPI","HOME+OFFICE", "TARGET <", "TIER1", "TIER2", "TIER3", "ENNOVA","Delta vs KPI"]
    hdoffice = ["INBOUND OFFICE", "TARGET <", "TIER1", "TIER2", "TIER3","ENNOVA","Delta vs KPI"]
    i = 1
    for title in header:
        i +=1
        if i != 9:
            ws.cell(column=i, row=headerdistance, value=title)
        else:
            i +=1
            ws.cell(column=i, row=headerdistance, value=title)

    x = 0
    for diz in opera_kpi:
        if diz == "FEO":
            x += 2
            i = 1
            for title in hdoffice:
                i +=1
                if i != 9:
                    ws.cell(column=i, row=headerdistance+x, value=title)
                else:
                    i +=1
                    ws.cell(column=i, row=headerdistance+x, value=title)
            col = ["B" + str(headerdistance+x) + ":H" + str(headerdistance+x)]
            style_range(ws, col[0], border=borderthin, fill=fillgray, font=fontboldblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
        elif diz == "BO":
            x = 0
        for i in range(0, len(opera_kpi[diz])):
            if diz != "BO":
                colnumb = 2
            else:
                colnumb = 10
            x += 1
            if diz == "BO":
                headerdistance = 6
                col = ["J" + str(headerdistance+x) + ":P" + str(headerdistance+x)]
            else:
                col = ["B" + str(headerdistance+x) + ":H" + str(headerdistance+x)]
            if (i % 2) == 0:
                if i == (len(opera_kpi[diz]) - 1):
                    for cells in col:
                        style_range(ws, cells, border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                for cells in col:
                    style_range(ws, cells, border=borderthinlat, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            else:
                if i == (len(opera_kpi[diz]) - 1):
                    for cells in col:
                        style_range(ws, cells, border=borderthinbot, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                for cells in col:
                    style_range(ws, cells, border=borderthinlat, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for data in opera_kpi[diz][i]:
                if data != "kpi_type":
                    ws.cell(column=colnumb, row=headerdistance+x, value=opera_kpi[diz][i][data])
                    init = str(ws.cell(column=colnumb, row=headerdistance+x)).find(".") + 1
                    if headerdistance+x > 9:
                        limit = init + 3
                    else:
                        limit = init + 2
                    if data == "kpo":
                        if opera_kpi[diz][i]["kpo"] < opera_kpi[diz][i]["tier3"]:
                            style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillblue, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif opera_kpi[diz][i]["kpo"] < opera_kpi[diz][i]["tier2"]:
                            style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillazure, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif opera_kpi[diz][i]["kpo"] < opera_kpi[diz][i]["tier1"]:
                            style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif opera_kpi[diz][i]["kpo"] < opera_kpi[diz][i]["target"]:
                            if opera_kpi[diz][i]["tier1"] == 0:
                                style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            else:
                                style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif opera_kpi[diz][i]["target"] == 0:
                            style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        else:
                            style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)

                    colnumb +=1

    ws.title = "Riepilogo"

    for i in range(0, len(indexdf)):
        wb = df_to_sheet(wb, df[indexdf[i]], indexdf[i])
    return wb


def c87_kpi_xls_template(wb, form, updatekpi, group, grcount):
    if grcount == 0:
        ws = wb.active
    else:
        ws = wb.create_sheet(group[grcount].BACINO)

    ws.sheet_view.zoomScale = 85

    headerdistance = 6
    reductedcol = ["A", "L"]

    try:
        for i in reductedcol:
            ws.column_dimensions[i].width = 2.78
        try:
            img = openpyxl.drawing.image.Image(r"C:\Apache\htdocs\CED_Cagliari\CED_Cagliari\static\app\img\excel_logo.png")
            img.anchor = "B2"
            ws.add_image(img)
        except FileNotFoundError:
            None

        ws["E2"] = "Estrazione di " + form["month"] + " " + form["year"]
        ws["B5"] = "Report Sintesi Volumi FE"
        ws["M5"] = "Report Sintesi Volumi BO"
        style_range(ws, "E2:H3", border=borderempty, fill=no_fill, font=fontboldblack, alignment=alignvertcenter, union=True, grid=False, interborder=False, width=None)

        headercol = ["B5:K5", "M5:V5"]
        for cells in headercol:
            style_range(ws, cells, border=borderthin, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=True, grid=False, interborder=False, width=None)

        headercol = ["B6", "C6:K6", "M6", "N6:V6"]
        for cells in headercol:
            if cells == "B6" or cells == "M6":
                width = 21.78
            else:
                width = 8.78
            style_range(ws, cells, border=borderthin, fill=fillgray, font=fontboldblack, alignment=aligncenter, union=False, grid=False, interborder=True, width=width)

        header = ["INBOUND HOME", "TARGET <", "TIER1", "TIER2", "TIER3", "ENNOVA","Delta vs KPI", "% Dec. Call", "% Malus", "Bonus/Malus", "HOME+OFFICE", "TARGET <", "TIER1", "TIER2", "TIER3", "ENNOVA", "Delta vs KPI", "% Dec. Call", "% Malus", "Bonus/Malus"]
        hdoffice = ["INBOUND OFFICE", "TARGET <", "TIER1", "TIER2", "TIER3","ENNOVA","Delta vs KPI", "% Dec. Call", "% Malus", "Bonus/Malus"]
        i = 1
        for title in header:
            i +=1
            if i != 12:
                ws.cell(column=i, row=headerdistance, value=title)
            else:
                i +=1
                ws.cell(column=i, row=headerdistance, value=title)

        x = 0
        for diz in updatekpi:
            if diz == "FEO":
                x += 2
                i = 1
                for title in hdoffice:
                    i +=1
                    if i != 12:
                        ws.cell(column=i, row=headerdistance+x, value=title)
                    else:
                        i +=1
                        ws.cell(column=i, row=headerdistance+x, value=title)
                col = ["B" + str(headerdistance+x) + ":K" + str(headerdistance+x)]
                style_range(ws, col[0], border=borderthin, fill=fillgray, font=fontboldblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            elif diz == "BO":
                x = 0
            for i in range(0, len(updatekpi[diz])):
                if diz != "BO":
                    colnumb = 2
                else:
                    colnumb = 13
                x += 1
                if diz == "BO":
                    headerdistance = 6
                    col = ["M" + str(headerdistance+x) + ":V" + str(headerdistance+x)]
                else:
                    col = ["B" + str(headerdistance+x) + ":K" + str(headerdistance+x)]
                if (i % 2) == 0:
                    if i == (len(updatekpi[diz]) - 1):
                        for cells in col:
                            style_range(ws, cells, border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                    for cells in col:
                        style_range(ws, cells, border=borderthinlat, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                else:
                    if i == (len(updatekpi[diz]) - 1):
                        for cells in col:
                            style_range(ws, cells, border=borderthinbot, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                    for cells in col:
                        style_range(ws, cells, border=borderthinlat, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                for data in updatekpi[diz][i]:
                    if data != "kpi_type":
                        ws.cell(column=colnumb, row=headerdistance+x, value=updatekpi[diz][i][data])
                        init = str(ws.cell(column=colnumb, row=headerdistance+x)).find(".") + 1
                        if headerdistance+x > 9:
                            limit = init + 3
                        else:
                            limit = init + 2
                        if data == "kpo":
                            if updatekpi[diz][i]["kpi_type"] == 1:
                                if updatekpi[diz][i]["kpo"] >= updatekpi[diz][i]["tier3"]:
                                    style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillblue, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                                elif updatekpi[diz][i]["kpo"] >= updatekpi[diz][i]["tier2"]:
                                    style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillazure, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                                elif updatekpi[diz][i]["kpo"] >= updatekpi[diz][i]["tier1"]:
                                    style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                                elif updatekpi[diz][i]["kpo"] >= updatekpi[diz][i]["target"]:
                                    style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillpink, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                                else:
                                    style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            else:
                                if updatekpi[diz][i]["kpo"] < updatekpi[diz][i]["tier3"]:
                                    style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillblue, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                                elif updatekpi[diz][i]["kpo"] < updatekpi[diz][i]["tier2"]:
                                    style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillazure, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                                elif updatekpi[diz][i]["kpo"] < updatekpi[diz][i]["tier1"]:
                                    style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                                elif updatekpi[diz][i]["kpo"] < updatekpi[diz][i]["target"]:
                                    if updatekpi[diz][i]["tier1"] == 0:
                                        style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                                    else:
                                        style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                                elif updatekpi[diz][i]["target"] == 0:
                                    style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                                else:
                                    style_range(ws, str(ws.cell(column=colnumb, row=headerdistance+x))[init:limit], border=borderempty, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)

                        colnumb +=1

    except Exception as error:
        print(type(error), error)


    if grcount == 0:
        ws.title = group[grcount].BACINO

    return wb


def recalls_xls_template(table, dfdown, datef, datel, htmlth, richmonth, richweek, tgt, contract):
    wb = Workbook()
    ws = wb.active

    ws.title = "Riepilogo"
    try:
        img = openpyxl.drawing.image.Image(r"C:\Apache\htdocs\CED_Cagliari\CED_Cagliari\static\app\img\excel_logo.png")
        img.anchor = "B2"
        ws.add_image(img)
    except FileNotFoundError:
        None
    ws['D2'] = "Estrazione Dati dal " + datef + " al " + datel
    title = "D2:H3"
    style_range(ws, title, border=borderempty, fill=no_fill, font=fontboldblack, alignment=aligncenter, union=True, grid=False, interborder=False, width=None)
    # Dati del Mese
    ws.column_dimensions["A"].width = 2.78
    ws["B5"] = "MENSILE"
    ws["J5"] = "MENSILE"
    ws["E5"] = "187 HOME"
    ws["M5"] = "191 OFFICE"

    widthcol = ["A", "D", "I", "L", "F", "N", "G", "O", "H", "P"]
    mutlimonth = 9
    for i in range(0, len(widthcol)):
        if i < 4:
            ws.column_dimensions[widthcol[i]].width = 2.78
        elif i > 3 and i < 6:
            ws.column_dimensions[widthcol[i]].width = 10.78
        elif i > 5 and i < 8:
            ws.column_dimensions[widthcol[i]].width = 19.78
        elif i > 7:
            ws.column_dimensions[widthcol[i]].width = 21.78


    headercol = ["B5:C5", "E5:H5", "J5:K5", "M5:P5"]
    for cells in headercol:
        style_range(ws, cells, border=borderthin, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=True, grid=False, interborder=False, width=None)

    headercol = ["B6:C6", "E6:H6", "J6:K6", "M6:P6"]
    for cells in headercol:
        style_range(ws, cells, border=borderthin, fill=fillgray, font=fontboldblack, alignment=aligncenter, union=False, grid=False, interborder=True, width=None)

    headerdistance = 6
    x = 0
    y = 0
    z = 0
    zz = 0
    for k in range(0, 17):
        if k < 2:
            ws.cell(column=k+2, row=headerdistance, value=htmlth["int"][k])
        elif k == 2 or k == 7:
            None
        elif k < 7 and k > 2:
            ws.cell(column=k+2, row=headerdistance, value=htmlth["inte"][x])
            x += 1
        elif k < 8:
            ws.cell(column=k+2, row=headerdistance, value=htmlth["inte"][y])
            y += 1
        elif k < 12 and k > 9:
            ws.cell(column=k, row=headerdistance, value=htmlth["int"][z])
            z += 1
        elif k > 12:
            ws.cell(column=k, row=headerdistance, value=htmlth["inte"][zz])
            zz += 1

    columlist187 = list(richmonth[table[0]][0][0])
    columlist191 = list(richmonth[table[1]][0][0])
    for u in range(0, len(richmonth[table[0]])):
        for i in range(0, len(richmonth[table[0]][u])):
            x = 2
            y = 2
            col = ["B" + str(headerdistance+i+1) + ":C" + str(headerdistance+i+1), "E" + str(headerdistance+i+1) + ":H" + str(headerdistance+i+1)]
            if i > 2:
                col = ["B" + str(mutlimonth+1) + ":C" + str(mutlimonth+1), "E" + str(mutlimonth+1) + ":H" + str(mutlimonth+1)]
                mutlimonth += 1
            if (i % 2) == 0:
                if i == (len(richmonth[table[0]][u]) - 1):
                    for cells in col:
                        style_range(ws, cells, border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                for cells in col:
                    style_range(ws, cells, border=borderthinlat, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            else:
                if i == (len(richmonth[table[0]][u]) - 1):
                    for cells in col:
                        style_range(ws, cells, border=borderthinbot, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                for cells in col:
                    style_range(ws, cells, border=borderthinlat, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for k in range(0, 7):
                if k < 2:
                    ws.cell(column=2, row=headerdistance+i+1, value=richmonth[table[0]][0][i][columlist187[0]])
                    ws.cell(column=3, row=headerdistance+i+1, value=richmonth[table[0]][0][i][columlist187[2]])
                    ws.column_dimensions["C"].width = 10.78
                    if i > 2:
                        ws.cell(column=2, row=mutlimonth, value=richmonth[table[0]][u][i][columlist187[0]])
                        ws.cell(column=3, row=mutlimonth, value=richmonth[table[0]][u][i][columlist187[2]])
                        ws.column_dimensions["B"].width = 10.78
                elif k == 2 or k == 7:
                    None
                elif k > 2 and k < 6:
                    x += 1
                    ws.cell(column=k+2, row=headerdistance+i+1, value=richmonth[table[0]][0][i][columlist187[x]])
                    valueperc = (str(richmonth[table[0]][0][i][columlist187[6]]) + " %")
                    ws["H"+str(headerdistance+i+1)] = valueperc
                    if i > 2:
                        ws.cell(column=k+2, row=mutlimonth, value=richmonth[table[0]][u][i][columlist187[x]])
                        valueperc = (str(richmonth[table[0]][u][i][columlist187[6]]) + " %")
                        ws["H"+str(mutlimonth)] = valueperc
                    cellfindf = str(ws.cell(column=8, row=headerdistance+i)).find(".") + 1
                    if headerdistance+i > 8:
                        cellfindl = cellfindf + 3
                    else:
                        cellfindl = cellfindf + 2
                    for g in range(0, 3):
                        if richmonth[table[0]][0][i][columlist187[2]] == richmonth[table[0]][0][g][columlist187[2]]:
                            if richmonth[table[0]][0][i][columlist187[6]] >= tgt["home"][g]["tgt"] and tgt["home"][g]["tgt"] != 0:
                                style_range(ws, str(ws.cell(column=8, row=headerdistance+i+1))[cellfindf:cellfindl], border=borderempty, fill=fillred, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            elif richmonth[table[0]][0][i][columlist187[6]] >= tgt["home"][g]["tier1"] and tgt["home"][g]["tier1"] != 0:
                                style_range(ws, str(ws.cell(column=8, row=headerdistance+i+1))[cellfindf:cellfindl], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            elif richmonth[table[0]][0][i][columlist187[6]] >= tgt["home"][g]["tier2"] and tgt["home"][g]["tier2"] != 0:
                                style_range(ws, str(ws.cell(column=8, row=headerdistance+i+1))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            elif richmonth[table[0]][0][i][columlist187[6]] >= tgt["home"][g]["tier3"] and tgt["home"][g]["tier3"] != 0:
                                style_range(ws, str(ws.cell(column=8, row=headerdistance+i+1))[cellfindf:cellfindl], border=borderempty, fill=fillazure, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            elif richmonth[table[0]][0][i][columlist187[6]] < tgt["home"][g]["tier3"] and tgt["home"][g]["tier3"] != 0:
                                style_range(ws, str(ws.cell(column=8, row=headerdistance+i+1))[cellfindf:cellfindl], border=borderempty, fill=fillblue, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            else:
                                style_range(ws, str(ws.cell(column=8, row=headerdistance+i+1))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
#office home
        multimonthw = 6
    for u in range(0, len(richmonth[table[1]])):
        for i in range(0, len(richmonth[table[1]][u])):
            multimonthw += 1
            col = ["J" + str(multimonthw) + ":K" + str(multimonthw), "M" + str(multimonthw) + ":P" + str(multimonthw)]
            if (i % 2) == 0:
                if i == (len(richmonth[table[1]][u]) - 1):
                    for cells in col:
                        style_range(ws, cells, border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                for cells in col:
                    style_range(ws, cells, border=borderthinlat, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            else:
                if i == (len(richmonth[table[1]][u]) - 1):
                    for cells in col:
                        style_range(ws, cells, border=borderthinbot, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                for cells in col:
                    style_range(ws, cells, border=borderthinlat, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for x in range(0, 7):
                if x < 2:
                    ws.cell(column=10, row=multimonthw, value=richmonth[table[1]][u][i][columlist191[0]])
                    ws.cell(column=11, row=multimonthw, value=richmonth[table[1]][u][i][columlist191[2]])
                    ws.column_dimensions["K"].width = 10.78
                    ws.column_dimensions["J"].width = 10.78
                elif x > 2 and x < 7:
                    ws.cell(column=x+10, row=multimonthw, value=richmonth[table[1]][u][i][columlist191[x]])
                    valueperc = (str(richmonth[table[1]][u][i][columlist191[6]]) + " %")
                    ws["P"+str(multimonthw)] = valueperc
                if contract[0].id == 1:
                    if richmonth[table[1]][u][i][columlist191[6]] >= tgt["office"][0]["tgt"] and tgt["office"][0]["tgt"] != 0:
                        style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillred, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                    elif richmonth[table[1]][u][i][columlist191[6]] >= tgt["office"][0]["tier1"] and tgt["office"][0]["tier1"] != 0:
                        style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                    elif richmonth[table[1]][u][i][columlist191[6]] >= tgt["office"][0]["tier2"] and tgt["office"][0]["tier2"] != 0:
                        style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                    elif richmonth[table[1]][u][i][columlist191[6]] >= tgt["office"][0]["tier3"] and tgt["office"][0]["tier3"] != 0:
                        style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillazure, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                    elif richmonth[table[1]][u][i][columlist191[6]] < tgt["office"][0]["tier3"] and tgt["office"][0]["tier3"] != 0:
                        style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillblue, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                    else:
                        style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                for g in range(0, 2):
                    if contract[0].id == 4:
                        cellfindf = str(ws.cell(column=8, row=multimonthw+i)).find(".") + 1
                        if int(multimonthw) > 9:
                            cellfindl = cellfindf + 3
                        else:
                            cellfindl = cellfindf + 2
                        if richmonth[table[1]][u][i][columlist191[6]] >= tgt["office"][g]["tgt"] and tgt["office"][g]["tgt"] != 0:
                            style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillred, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richmonth[table[1]][u][i][columlist191[6]] >= tgt["office"][g]["tier1"] and tgt["office"][g]["tier1"] != 0:
                            style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richmonth[table[1]][u][i][columlist191[6]] >= tgt["office"][g]["tier2"] and tgt["office"][g]["tier2"] != 0:
                            style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richmonth[table[1]][u][i][columlist191[6]] >= tgt["office"][g]["tier3"] and tgt["office"][g]["tier3"] != 0:
                            style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillazure, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richmonth[table[1]][u][i][columlist191[6]] < tgt["office"][g]["tier3"] and tgt["office"][g]["tier3"] != 0:
                            style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillblue, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        else:
                            style_range(ws, str(ws.cell(column=16, row=multimonthw))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)

    columlistweek187 = list(richweek[table[0]][0][0])
    columlistweek191 = list(richweek[table[1]][0][0])
    x = 0
    y = 0
    z = 0
    zz = 0
    officerow = (7 + int(len(richmonth[table[0]][0])))

    ws["B" + str(officerow + 1)] = "SETTIMANALE"
    ws["J" + str(officerow + 1)] = "SETTIMANALE"
    ws["E" + str(officerow + 1)] = "187 HOME"
    ws["M" + str(officerow + 1)] = "191 OFFICE"
    for k in range(0, 17):
        if k < 2:
            ws.cell(column=k+2, row=officerow+2, value=htmlth["int"][k])
        elif k == 2 or k == 7:
            None
        elif k < 7 and k > 2:
            ws.cell(column=k+2, row=officerow+2, value=htmlth["inte"][x])
            x += 1
        elif k < 8:
            ws.cell(column=k+2, row=officerow+2, value=htmlth["inte"][y])
            y += 1
        elif k < 12 and k > 9:
            ws.cell(column=k, row=officerow+2, value=htmlth["int"][z])
            z += 1
        elif k > 12:
            ws.cell(column=k, row=officerow+2, value=htmlth["inte"][zz])
            zz += 1

    headercol = ["B" + str(officerow+1) + ":" + "C" + str(officerow+1), "E" + str(officerow+1) + ":" + "H" + str(officerow+1), "J" + str(officerow+1) + ":" + "K" + str(officerow+1), "M" + str(officerow+1) + ":" + "P" + str(officerow+1)]
    for cells in headercol:
        style_range(ws, cells, border=borderthin, fill=fillred, font=fontboldwhite, alignment=aligncenter, union=True, grid=False, interborder=False, width=None)
    headercol = ["B" + str(officerow+2) + ":" + "C" + str(officerow+2), "E" + str(officerow+2) + ":" + "H" + str(officerow+2), "J" + str(officerow+2) + ":" + "K" + str(officerow+2), "M" + str(officerow+2) + ":" + "P" + str(officerow+2)]
    for cells in headercol:
        style_range(ws, cells, border=borderthin, fill=fillgray, font=fontboldblack, alignment=aligncenter, union=False, grid=False, interborder=True, width=None)

    for i in range(0, len(richweek[table[0]][0])):
        x = 2
        y = 2
        col = ["B" + str(officerow+i+3) + ":C" + str(officerow+i+3), "E" + str(officerow+i+3) + ":H" + str(officerow+i+3)]
        if (i % 2) == 0:
            if i == (len(richweek[table[0]][0]) - 1):
                for cells in col:
                    style_range(ws, cells, border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for cells in col:
                style_range(ws, cells, border=borderthinlat, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
        else:
            if i == (len(richweek[table[0]][0]) - 1):
                for cells in col:
                    style_range(ws, cells, border=borderthinbot, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for cells in col:
                style_range(ws, cells, border=borderthinlat, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
        for k in range(0, 7):
            if k < 2:
                ws.cell(column=2, row=officerow+i+3, value=richweek[table[0]][0][i][columlistweek187[0]])
                ws.cell(column=3, row=officerow+i+3, value=richweek[table[0]][0][i][columlistweek187[2]])
                ws.column_dimensions["C"].width = 10.78
            elif k == 2 or k == 7:
                None
            elif k > 2 and k < 7:
                x += 1
                ws.cell(column=k+2, row=officerow+i+3, value=richweek[table[0]][0][i][columlistweek187[x]])
                valueperc = (str(richweek[table[0]][0][i][columlistweek187[6]]) + " %")
                ws["H"+str(officerow+i+3)] = valueperc
                cellfindf = str(ws.cell(column=8, row=officerow+i+3)).find(".") + 1
                cellfindl = cellfindf + 3
                for g in range(0, 3):
                    if richweek[table[0]][0][i][columlistweek187[2]] == richweek[table[0]][0][g][columlistweek187[2]]:
                        if richweek[table[0]][0][i][columlistweek187[6]] >= tgt["home"][g]["tgt"] and tgt["home"][g]["tgt"] != 0:
                            style_range(ws, str(ws.cell(column=8, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillred, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richweek[table[0]][0][i][columlistweek187[6]] >= tgt["home"][g]["tier1"] and tgt["home"][g]["tier1"] != 0:
                            style_range(ws, str(ws.cell(column=8, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richweek[table[0]][0][i][columlistweek187[6]] >= tgt["home"][g]["tier2"] and tgt["home"][g]["tier2"] != 0:
                            style_range(ws, str(ws.cell(column=8, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richweek[table[0]][0][i][columlistweek187[6]] >= tgt["home"][g]["tier3"] and tgt["home"][g]["tier3"] != 0:
                            style_range(ws, str(ws.cell(column=8, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillazure, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richweek[table[0]][0][i][columlistweek187[6]] < tgt["home"][g]["tier3"] and tgt["home"][g]["tier3"] != 0:
                            style_range(ws, str(ws.cell(column=8, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillblue, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        else:
                            style_range(ws, str(ws.cell(column=8, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
#office week
    for i in range(0, len(richweek[table[1]][0])):
        t = 2
        col = ["J" + str(officerow+i+3) + ":K" + str(officerow+i+3), "M" + str(officerow+i+3) + ":P" + str(officerow+i+3)]
        if (i % 2) == 0:
            if i == (len(richweek[table[1]][0]) - 1):
                for cells in col:
                    style_range(ws, cells, border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for cells in col:
                style_range(ws, cells, border=borderthinlat, fill=None, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
        else:
            if i == (len(richweek[table[1]][0]) - 1):
                for cells in col:
                    style_range(ws, cells, border=borderthinbot, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
            for cells in col:
                style_range(ws, cells, border=borderthinlat, fill=filllighgray, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
        for x in range(0, 7):
            if x < 2:
                ws.cell(column=10, row=officerow+i+3, value=richweek[table[1]][0][i][columlist191[0]])
                ws.cell(column=11, row=officerow+i+3, value=richweek[table[1]][0][i][columlist191[2]])
                ws.column_dimensions["K"].width = 10.78
            elif x == 2 or x == 7:
                None
            elif x > 2 and x <= 6:
                t += 1
                ws.cell(column=x+10, row=officerow+i+3, value=richweek[table[1]][0][i][columlist191[t]])
                valueperc = (str(richweek[table[1]][0][i][columlist191[6]]) + " %")
                ws["P"+str(officerow+i+3)] = valueperc
                cellfindf = str(ws.cell(column=16, row=officerow+i+3)).find(".") + 1
                cellfindl = cellfindf + 3
                if contract[0].id == 1:
                    if richweek[table[1]][0][i][columlist191[2]] == richweek[table[1]][0][0][columlist191[2]]:
                        if richweek[table[1]][0][i][columlist191[6]] >= tgt["office"][0]["tgt"] and tgt["office"][0]["tgt"] != 0:
                            style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillred, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richweek[table[1]][0][i][columlistweek187[6]] >= tgt["office"][0]["tier1"] and tgt["office"][0]["tier1"] != 0:
                            style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richweek[table[1]][0][i][columlistweek187[6]] >= tgt["office"][0]["tier2"] and tgt["office"][0]["tier2"] != 0:
                            style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richweek[table[1]][0][i][columlistweek187[6]] >= tgt["office"][0]["tier3"] and tgt["office"][0]["tier3"] != 0:
                            style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillazure, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        elif richweek[table[1]][0][i][columlistweek187[6]] < tgt["office"][0]["tier3"] and tgt["office"][0]["tier3"] != 0:
                            style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillblue, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                        else:
                            style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                for g in range(0, 2):
                    if contract[0].id == 4:
                        if richweek[table[1]][0][i][columlist191[2]] == richweek[table[1]][0][g][columlist191[2]]:
                            if richweek[table[1]][0][i][columlist191[6]] >= tgt["office"][g]["tgt"] and tgt["office"][g]["tgt"] != 0:
                                style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillred, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            elif richweek[table[1]][0][i][columlistweek187[6]] >= tgt["office"][g]["tier1"] and tgt["office"][g]["tier1"] != 0:
                                style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillpink, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            elif richweek[table[1]][0][i][columlistweek187[6]] >= tgt["office"][g]["tier2"] and tgt["office"][g]["tier2"] != 0:
                                style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            elif richweek[table[1]][0][i][columlistweek187[6]] >= tgt["office"][g]["tier3"] and tgt["office"][g]["tier3"] != 0:
                                style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillazure, font=fontwhite, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            elif richweek[table[1]][0][i][columlistweek187[6]] < tgt["office"][g]["tier3"] and tgt["office"][g]["tier3"] != 0:
                                style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillblue, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)
                            else:
                                style_range(ws, str(ws.cell(column=16, row=officerow+i+3))[cellfindf:cellfindl], border=borderempty, fill=fillgreen, font=fontblack, alignment=aligncenter, union=None, grid=False, interborder=True, width=None)

    for i in range(0, len(richmonth[table[0]][0]), 3):
        style_range(ws, 'B'+str(headerdistance+1+i)+':'+'B'+str(headerdistance+3+i), border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=True, grid=False, interborder=True, width=None)
    for i in range(0, len(richweek[table[0]][0]), 3):
        style_range(ws, 'B'+str(officerow+3+i)+':'+'B'+str(officerow+5+i), border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=True, grid=False, interborder=True, width=None)
    for i in range(0, len(richmonth[table[1]][0]), 2):
        if contract[0].id == 4:
            style_range(ws, 'J'+str(headerdistance+1+i)+':'+'J'+str(headerdistance+2+i), border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=True, grid=False, interborder=True, width=None)
    for i in range(0, len(richweek[table[1]][0]), 2):
        if contract[0].id == 4:
            style_range(ws, 'J'+str(officerow+3+i)+':'+'J'+str(officerow+4+i), border=borderthinbot, fill=None, font=fontblack, alignment=aligncenter, union=True, grid=False, interborder=True, width=None)
    wb = df_to_sheet(wb, dfdown[0], table[0][4:], index=False)
    wb = df_to_sheet(wb, dfdown[1], table[1][4:], index=False)

    return wb
