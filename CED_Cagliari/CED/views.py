import datetime as dt
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook, load_workbook
import random
import os


import pandas as pd
import decimal
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import connection
from django.http import Http404, HttpResponse, HttpResponseRedirect
from django.shortcuts import render

import CED.functions.dataframeoperations as dfo
import CED.functions.viewsoperations as vo
import CED.functions.fileeditor_xls as fe
import CED.templates.Mail.mail_template as mt
import CED.var.links_paths as lp
import CED.var.query as qu
import CED.var.targets as tz
from CED.models import c87_group, month_tr, c87_KPI_KPO_obtained, tools_credentials, c87_ASC, c87_ibia_rows, \
    tkt_money_value, days_tr, order_tracker, discount_tracker, sites_tracker, contract_tracker, kpi_tracker, \
    supervisor_tracker, festivity_tr, db_update_tracker


def login_view(request):
    return render(request, "registration/login.html")


def checklogin(request):
    if request.method != 'POST':
        raise Http404('Only POSTs are allowed')
    try:
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_active:
                request.session.set_expiry(7400)
                login(request, user)
                return index(request)
        else:
            messages.error(request, "Nome Utente o Password Errata.")
            return login_view(request)
    except User.DoesNotExist:
            messages.error(request, "Nome Utente o Password Errata.")
            return login_view(request)


@login_required
def logout_view(request):
    logout(request)


@login_required
def index(request):
    context = []
    dates = dict()
    dbetiquette = ["IVR", "OpeRA", "IBIA_SOS_EASY"]
    for ett in dbetiquette:
        dates[ett] = db_update_tracker.objects.raw("SELECT *, max(date_hour) as dbup FROM ced_db_update_tracker WHERE db_name = '%s'" % ett)[0].dbup
    context = {"dates": dates}
    return render(request, "CED/index.html", context)


@login_required
def download(request):
    try:
        status = request.POST["status"]
    except Exception as error:
        print(type(error), error)
    if status == "1":
        try:
            wb = load_workbook(request.session["dir"] + request.session["filename"])
            os.remove(request.session["dir"] + request.session["filename"])
            os.rmdir(request.session["dir"])
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="%s"' % request.session["filename"]
            wb.save(response)
            return response
        except FileNotFoundError:
            error = "File non trovato - L'errore può essere dovuto all'utilizzo dei tasti Avanti/Indietro del browser o al fatto che il file sia già stato scaricato"
            context = {"error": error}
            return render(request, 'CED/download.html', context)
    elif status == "2":
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required
def ivr(request):
    mdict = dict()
    form = dict()
    errors = dict()
    groups = []
    ivrsmonth = dict()
    ivrsweek = dict()
    target = dict()
    df = pd.DataFrame()
    firstdf = pd.DataFrame()
    date_from = 0
    date_to = 0
    ivrtable = ""
    info = ""
    tgt = dict()

    table = ["ced_c87_ivr_187", "ced_c87_ivr_191"]
    htmlth = {"int": ["Periodo", "Gruppo"],
              "187new": ["IVR Validi", "IVR NV", "Media D4", "Media D7", "Semplici", "Medio Comp.", "Complesse"],
              "191new": ["IVR Validi", "IVR NV", "Media D4", "Media D7", "Medio Comp.", "Complesse"],
              "quant": ["Periodo", "Gruppo", "IVR Validi", "IVR NV", "Voto_1", "Voto_2", "Voto_3", "Voto_4", "Voto_5", "Voto_6",
                        "Voto_7", "Voto_8", "Voto_9", "Voto_10"]}

    for i in range(0, len(table)):
        ivrsmonth[table[i]] = []
        ivrsweek[table[i]] = []

    groups = c87_group.objects.raw("SELECT * FROM ced_c87_group WHERE substring(BACINO, 1, 9) = 'ENNOVA_CA' OR "
                                   "BACINO = 'ENNOVA' ORDER BY BACINO")
    months = month_tr.objects.raw("SELECT name, number FROM ced_month_tr ORDER BY number")
    order = order_tracker.objects.raw("SELECT * FROM ced_order_tracker WHERE pafdesc = 'C87'")

    try:
        form["year"] = int(request.POST["year"])
        ivrtable = request.POST["ivrtable"]
        form["group"] = request.POST["group"]
        try:
            form["month"] = request.POST["month"]
            form["date_from"] = request.POST["date_from"]
            form["date_to"] = request.POST["date_to"]
            monthquery = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE number >= '%s' "
                                              "AND number <= '%s'" % (form["date_from"][5:7], form["date_to"][5:7]))
            date_from = dt.datetime.strptime(form["date_from"], "%Y-%m-%d")
            date_to = dt.datetime.strptime(form["date_to"], "%Y-%m-%d")
        except:
            form["month"] = request.POST["month"]
            monthquery = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE name = '%s'" % form["month"])
        if (form["month"] == "" and (form["date_from"] == "" or form["date_to"] == "")):
            errors["empty"] = "Inserisci una Data"
        elif date_to < date_from:
            errors["date"] = "Data Futura o non Congruente"
        elif (form["month"] != "") and (int(str(form["year"]) + monthquery[0].number) > int(str(dt.datetime.now().year) + str(dt.datetime.now().month).zfill(2))):
            errors["date"] = "Data Futura o non Congruente"
        elif errors == {}:
            if form["month"] != "":
                sqldate = dt.datetime.strptime(str(form["year"]) + monthquery[0].number + "01", "%Y%m%d")
                contract = contract_tracker.objects.raw("""SELECT * FROM ced_contract_tracker WHERE ced_contract_tracker.order_id = '%s'
                                                        AND fromdate <= '%s' AND todate >= '%s' ORDER BY id""" % (order[0].id, sqldate, sqldate))
            else:
                contract = contract_tracker.objects.raw("""SELECT * FROM ced_contract_tracker WHERE ced_contract_tracker.order_id = '%s'
                                                        AND fromdate <= '%s' AND todate >= '%s' ORDER BY id""" % (order[0].id, date_from, date_to))

            if len(list(contract)) == 0:
                errors["date"] = "Nessuno o più contratti sono attivi nelle date scelte, scegli un range differente"
            else:
                target = kpi_tracker.objects.raw("""SELECT id, `name`, target, tier1, tier2, tier3, bonustarget, bonustier1,
                                                    bonustier2, bonustier3, bonustier4 FROM ced_kpi_tracker
                                                   WHERE contract_id = '%s' AND tool = 'PACO IVR' ORDER BY sortorder""" % (contract[0].id))
                tgt["home"] = list()
                tgt["office"] = list()
                for i in range(0, 4):
                    tgt["home"].append(dict())
                    if contract[0].id == 1:
                        if i == 0:
                            tgt["home"][0]["tgt"] = target[0].target
                            tgt["home"][0]["tier1"] = target[0].tier1
                            tgt["home"][0]["tier2"] = target[0].tier2
                            tgt["home"][0]["tier3"] = target[0].tier3
                        else:
                            tgt["home"][i]["tgt"] = 0
                            tgt["home"][i]["tier1"] = 0
                            tgt["home"][i]["tier2"] = 0
                            tgt["home"][i]["tier3"] = 0
                    elif contract[0].id == 4:
                        if i == 0:
                            tgt["home"][0]["tgt"] = 0
                            tgt["home"][0]["tier1"] = 0
                            tgt["home"][0]["tier2"] = 0
                            tgt["home"][0]["tier3"] = 0
                        else:
                            tgt["home"][i]["tgt"] = target[i-1].target
                            tgt["home"][i]["tier1"] = target[i-1].tier1
                            tgt["home"][i]["tier2"] = target[i-1].tier2
                            tgt["home"][i]["tier3"] = target[i-1].tier3
                    j = i+1

                for i in range(0, 3):
                    tgt["office"].append(dict())
                    if contract[0].id == 1:
                        if i == 0:
                            tgt["office"][0]["tgt"] = target[1].target
                            tgt["office"][0]["tier1"] = target[1].tier1
                            tgt["office"][0]["tier2"] = target[1].tier2
                            tgt["office"][0]["tier3"] = target[1].tier3
                        else:
                            tgt["office"][i]["tgt"] = 0
                            tgt["office"][i]["tier1"] = 0
                            tgt["office"][i]["tier2"] = 0
                            tgt["office"][i]["tier3"] = 0
                    elif contract[0].id == 4:
                        if i == 0:
                            tgt["office"][0]["tgt"] = 0
                            tgt["office"][0]["tier1"] = 0
                            tgt["office"][0]["tier2"] = 0
                            tgt["office"][0]["tier3"] = 0
                        else:
                            tgt["office"][i]["tgt"] = target[i+2].target
                            tgt["office"][i]["tier1"] = target[i+2].tier1
                            tgt["office"][i]["tier2"] = target[i+2].tier2
                            tgt["office"][i]["tier3"] = target[i+2].tier3

            for i in range(0, len(table)):
                date = []
                years = []
                if form["month"] != "":
                    fweek = int(dt.date(form["year"], int(monthquery[0].number), 1).strftime("%W"))
                    try:
                        lweek = int((dt.date(form["year"], int(monthquery[0].number) + 1, 1)-dt.timedelta(1)).strftime("%W"))
                    except ValueError:
                        lweek = 1
                    if form["group"] == "ENNOVA":
                        df = pd.read_sql("SELECT * FROM %s WHERE substring(MODULO,1, 6) = '%s' AND substring(DATA_INT,4, 7) = '%s'"
                                        % (table[i], form["group"], (monthquery[0].number+"/"+str(form["year"]))), connection)
                    else:
                        df = pd.read_sql("SELECT * FROM %s WHERE MODULO = '%s' AND substring(DATA_INT,4, 7) = '%s'"
                                        % (table[i], form["group"], (monthquery[0].number+"/"+str(form["year"]))), connection)
                    date.append(monthquery[0].number + "/" + str(form["year"]))
                    years.append(str(form["year"]))
                else:
                    if int(form["date_from"][:4]) == int(form["date_to"][:4]):
                        monthquery = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE number >= '%s' AND number <= '%s' ORDER BY number"
                                                        % (form["date_from"][5:7], form["date_to"][5:7]))
                        for month in monthquery:
                            date.append(month.number + "/" + form["date_from"][:4])
                        years.append(form["date_from"][:4])
                    else:
                        monthquery1 = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE number >= '%s' AND number <= '12' ORDER BY number"
                                                        % (form["date_from"][5:7]))
                        monthquery2 = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE number >= '01' AND number <= '%s' ORDER BY number"
                                                        % (form["date_to"][5:7]))
                        for month in monthquery1:
                            date.append(month.number + "/" + form["date_from"][:4])
                        for month in monthquery2:
                            date.append(month.number + "/" + form["date_to"][:4])
                        years.append(form["date_from"][:4])
                        years.append(form["date_to"][:4])
                    mcount = len(date)
                    df = pd.read_sql("SELECT * FROM " + table[i] + " WHERE substring(MODULO,1, 6) = '" + form["group"] +
                                     "' AND str_to_date(DATA_INT,'%d/%m/%Y') >= '" + form["date_from"] + "' AND str_to_date(DATA_INT,'%d/%m/%Y') <= '"
                                     + form["date_to"] + "'", connection)
                    fweek = int(date_from.strftime("%W"))
                    lweek = int(date_to.strftime("%W"))
                if df.empty:
                    errors["empty"] = "Non ci sono IVR nel periodo selezionato"
                else:
                    info = "NB. Selezionando un mese o un range di date che non parta dal lunedì e non finisca Domenica, la prima e l'ultima settimana potrebbero avere giorni mancanti."
                    ivrsmonth[table[i]].append(dfo.ivrmonth(df, table[i], date, form["group"]).copy())
                    ivrsweek[table[i]].append(dfo.ivrweek(df, table[i], fweek, lweek, years, form["group"]).copy())
                    if ivrtable == "1" and i == 0:
                        firstdf = df
            if ivrtable == "1":
                if form["month"] != "":
                    datefrom = str(form["year"]) + "-" + str(monthquery[0].number) + "-01"
                    dateto = str(form["year"]) + "-" + str(monthquery[0].number) + "-" + str(monthquery[0].days)
                else:
                    datefrom = form["date_from"]
                    dateto = form["date_to"]
                wb = fe.c87_ivr_xls_template(table, ivrsmonth, ivrsweek, htmlth, datefrom, dateto, firstdf, df, tgt)
                filename = "Dettaglio ivr 187 191 " + form["group"] + ".xlsx"
                try:
                    key = '%016x' % random.getrandbits(64)
                    os.makedirs("./CED/downloads/" + key)
                    wb.save("./CED/downloads/" + key + "/" + filename)
                    request.session["filename"] = filename
                    request.session["dir"] = "./CED/downloads/" + key + "/"
                    return render(request, 'CED/download.html')
                except Exception as error:
                    print(type(error), error)
    except Exception as error:
        print(type(error), error)
    context = {0: {'groups': groups}, 1: {'ivrsmonth': ivrsmonth},
               2: {'ivrsweek': ivrsweek}, 3: {'months': months},
               6: {'htmlth': htmlth}, 7: {'form': form},
               8: {'ivrtable': ivrtable},
               9: {'errors': errors}, 10: {'target': tgt},
               4: {'info': info}}
    return render(request, "CED/ivr.html", context)


@login_required
def opera(request):
    form = dict()
    errors = dict()
    dfdict = dict()
    indexdfdict = ["FEH", "FEO", "BOH", "BOO"]
    kpo = []

    opera_kpi = dict()
    opera_kpi["FEH"] = list()
    opera_kpi["FEO"] = list()
    opera_kpi["BO"] = list()

    groups = c87_group.objects.raw("SELECT * FROM ced_c87_group WHERE substring(BACINO, 1, 9) = 'ENNOVA_CA' OR BACINO = 'ENNOVA' ORDER BY BACINO")
    months = month_tr.objects.raw("SELECT name, number FROM ced_month_tr ORDER BY number")
    order = order_tracker.objects.raw("SELECT * FROM ced_order_tracker WHERE pafdesc = 'C87'")

    table = ["ced_opera_sintesi_volumi_fe_home", "ced_opera_sintesi_volumi_fe_office",
             "ced_opera_sintesi_volumi_bo_home", "ced_opera_sintesi_volumi_bo_office"]
    try:
        form["year"] = request.POST["year"]
        form["month"] = request.POST["month"]
        form["group"] = request.POST["group"]
        form["tableopera"] = request.POST["tableopera"]
        month = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE name = '%s'" % form["month"])
        group = c87_group.objects.raw("SELECT * FROM ced_c87_group WHERE BACINO = '%s'" % form["group"])
        sqldate = dt.datetime.strptime(form["year"] + month[0].number + "01", "%Y%m%d")
        contract = contract_tracker.objects.raw("""SELECT * FROM ced_contract_tracker WHERE ced_contract_tracker.order_id = '%s'
                                               AND fromdate <= '%s' AND todate >= '%s' ORDER BY id""" % (order[0].id, sqldate, sqldate))
        target = kpi_tracker.objects.raw("""SELECT id, `name`, target, tier1, tier2, tier3, bonustarget, bonustier1,
                                               bonustier2, bonustier3, bonustier4 FROM ced_kpi_tracker
                                               WHERE contract_id = '%s' AND tool = 'OPERA' ORDER BY sortorder""" % (contract[0].id))
        date = month[0].number + form["year"][2:]
        if (form["month"] == "" or form["year"] == ""):
            errors["empty"] = "Inserisci una Data"
        elif (int((form["year"]+month[0].number)) > int((str(dt.datetime.now().year)+str(dt.datetime.now().month).zfill(2)))):
            errors["date"] = "Data Futura o non Congruente"
        else:
            for i in range(0, 4):
                querykpi = qu.opera_month_kpi(table[i], date, group[0].ID)
                df = pd.read_sql(querykpi, connection, index_col="Keys_SERVIZIO_id")
                columns = df.columns.tolist()
                if df.empty:
                    errors["empty"] = "Non ci sono dati nel periodo selezionato"
                elif i == 0:
                    querykpi = qu.opera_month_kpi_dl(table[i], date, group[0].ID)
                    dfdict[indexdfdict[i]] = pd.read_sql(querykpi, connection)
                    if contract[0].id == 1:
                        kpo = dfo.KpiFEHome2017(df, columns)
                    elif contract[0].id == 4:
                        kpo = dfo.KpiFEHome2018(df, columns)
                    for i in range(0, len(kpo)):
                        opera_kpi["FEH"].append(dict())
                        opera_kpi["FEH"][i]["name"] = target[i].name
                        opera_kpi["FEH"][i]["target"] = target[i].target
                        opera_kpi["FEH"][i]["tier1"] = target[i].tier1
                        opera_kpi["FEH"][i]["tier2"] = target[i].tier2
                        opera_kpi["FEH"][i]["tier3"] = target[i].tier3
                        opera_kpi["FEH"][i]["kpo"] = kpo[i]
                        if opera_kpi["FEH"][i]["target"] == 0:
                            opera_kpi["FEH"][i]["delta"] = opera_kpi["FEH"][i]["tier1"] - opera_kpi["FEH"][i]["kpo"]
                        else:
                            opera_kpi["FEH"][i]["delta"] = opera_kpi["FEH"][i]["target"] - opera_kpi["FEH"][i]["kpo"]
                    j = i+1

                elif i == 1:
                    querykpi = qu.opera_month_kpi_dl(table[i], date, group[0].ID)
                    dfdict[indexdfdict[i]] = pd.read_sql(querykpi, connection)
                    if contract[0].id == 1:
                        kpo = dfo.KpiFEOffice2017(df, columns)
                    elif contract[0].id == 4:
                        kpo = dfo.KpiFEOffice2018(df, columns)
                    for i in range(0, len(kpo)):
                        opera_kpi["FEO"].append(dict())
                        opera_kpi["FEO"][i]["name"] = target[i+j].name
                        opera_kpi["FEO"][i]["target"] = target[i].target
                        opera_kpi["FEO"][i]["tier1"] = target[i].tier1
                        opera_kpi["FEO"][i]["tier2"] = target[i].tier2
                        opera_kpi["FEO"][i]["tier3"] = target[i].tier3
                        opera_kpi["FEO"][i]["kpo"] = kpo[i]
                        if opera_kpi["FEO"][i]["target"] == 0:
                            opera_kpi["FEO"][i]["delta"] = opera_kpi["FEO"][i]["tier1"] - opera_kpi["FEO"][i]["kpo"]
                        else:
                            opera_kpi["FEO"][i]["delta"] = opera_kpi["FEO"][i]["target"] - opera_kpi["FEO"][i]["kpo"]
                    j += i+1

                elif i == 2:
                    df2 = pd.read_sql(querykpi, connection, index_col="Keys_SERVIZIO_id")
                    querykpi = qu.opera_month_kpi_dl(table[i], date, group[0].ID)
                    dfdict[indexdfdict[i]] = pd.read_sql(querykpi, connection)
                elif i == 3:
                    querykpi = qu.opera_month_kpi_dl(table[i], date, group[0].ID)
                    dfdict[indexdfdict[i]] = pd.read_sql(querykpi, connection)
                    if contract[0].id == 1:
                        kpo = dfo.KpiBOHoOf2017(df, df2, columns)
                    elif contract[0].id == 4:
                        kpo = dfo.KpiBOHoOf2018(df, df2, columns)
                    for i in range(0, len(kpo)):
                        opera_kpi["BO"].append(dict())
                        opera_kpi["BO"][i]["name"] = target[i+j].name
                        opera_kpi["BO"][i]["target"] = target[i].target
                        opera_kpi["BO"][i]["tier1"] = target[i].tier1
                        opera_kpi["BO"][i]["tier2"] = target[i].tier2
                        opera_kpi["BO"][i]["tier3"] = target[i].tier3
                        opera_kpi["BO"][i]["kpo"] = kpo[i]
                        if opera_kpi["BO"][i]["target"] == 0:
                            opera_kpi["BO"][i]["delta"] = opera_kpi["BO"][i]["tier1"] - opera_kpi["BO"][i]["kpo"]
                        else:
                            opera_kpi["BO"][i]["delta"] = opera_kpi["BO"][i]["target"] - opera_kpi["BO"][i]["kpo"]

        if form["tableopera"] == "1":
            for i in range(0, len(indexdfdict)):
                dfdict[indexdfdict[i]].sort_values(by=["Keys_SERVIZIO_id"], ascending=False, inplace=True)
                dfdict[indexdfdict[i]].replace(dfdict[indexdfdict[i]]["Keys_BACINO_id"][0], dfdict[indexdfdict[i]]["BACINO"][0], inplace=True)
                for j in range(0, len(dfdict[indexdfdict[i]]["Keys_SERVIZIO_id"].tolist())):
                    dfdict[indexdfdict[i]].replace(dfdict[indexdfdict[i]]["Keys_SERVIZIO_id"][j], dfdict[indexdfdict[i]]["SERVICE"][j], inplace=True)
                dfdict[indexdfdict[i]].drop(["ID", "BACINO", "SERVICE"], axis=1, inplace=True)
            wb = fe.c87_opera_xls_template(form, opera_kpi, target, dfdict, indexdfdict)
            filename = "Dettaglio OpeRA 187 191 " + form["group"] + ".xlsx"
            try:
                key = '%016x' % random.getrandbits(64)
                os.makedirs("./CED/downloads/" + key)
                wb.save("./CED/downloads/" + key + "/" + filename)
                request.session["filename"] = filename
                request.session["dir"] = "./CED/downloads/" + key + "/"
                return render(request, 'CED/download.html')
            except Exception as error:
                print(type(error), error)
    except:
        None
    context = {0: {'groups': groups},
               2: {'opera_kpi': opera_kpi},
               4: {'months': months},

               7: {'form': form}, 8: {'errors': errors}}
    return render(request, "CED/opera.html", context)


@login_required
def richiamate(request):
    mdict = dict()
    form = dict()
    errors = dict()
    target = dict()
    richtable = []
    dflistdown = []
    richmonth = dict()
    richweek = dict()
    df = pd.DataFrame()
    contract = ""
    info = ""
    tgt = dict()
    date_from = 0
    date_to = 0
    date = 0
    datef = ""
    datel = ""
    monthcheck = "00"
    table = ["ced_paco_richiamate_fe_187", "ced_paco_richiamate_fe_191"]
    htmlth = {"int": ["Periodo", "Tipologia"],
              "inte": ["Chiamate", "Abbattute", "Richiamate 3 Giorni", "Richiamate 3 Giorni %"]}

    for i in range(0, len(table)):
        richmonth[table[i]] = []
        richweek[table[i]] = []

    groups = c87_group.objects.raw("SELECT ID, BACINO FROM ced_c87_group WHERE substring(BACINO, 1, 9) = 'ENNOVA_CA' OR BACINO = 'ENNOVA' ORDER BY BACINO")
    months = month_tr.objects.raw("SELECT name, number FROM ced_month_tr ORDER BY number")
    order = order_tracker.objects.raw("SELECT * FROM ced_order_tracker WHERE pafdesc = 'C87'")

    try:
        form["Year"] = int(request.POST["Year"])
        form["group"] = request.POST["group"]
        richtable = request.POST["btn"]
        try:
            form["month"] = request.POST["month"]
            form["date_from"] = request.POST["date_from"]
            form["date_to"] = request.POST["date_to"]
            date_from = dt.datetime.strptime(form["date_from"], "%Y-%m-%d")
            date_to = dt.datetime.strptime(form["date_to"], "%Y-%m-%d")
            month = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE number >= '%s' "
                                              "AND number <= '%s'" % (form["date_from"][5:7], form["date_to"][5:7]))
        except:
            form["month"] = request.POST["month"]
            month = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE name = '%s'" % form["month"])
            sqldate = dt.datetime.strptime(str(form["Year"]) + month[0].number + "01", "%Y%m%d")
        group = c87_group.objects.raw("SELECT ID FROM ced_c87_group WHERE BACINO = '%s'" % form["group"])
        if form["month"] != "":
            contract = contract_tracker.objects.raw("""SELECT * FROM ced_contract_tracker WHERE ced_contract_tracker.order_id = '%s'
                                                   AND fromdate <= '%s' AND todate >= '%s' ORDER BY id""" % (order[0].id, sqldate, sqldate))
        else:
            contract = contract_tracker.objects.raw("""SELECT * FROM ced_contract_tracker WHERE ced_contract_tracker.order_id = '%s'
                                                   AND fromdate <= '%s' AND todate >= '%s' ORDER BY id""" % (order[0].id, date_from, date_to))
        if len(list(contract)) == 0:
            errors["date"] = "Nessuno o più contratti sono attivi nelle date scelte, scegli un range differente"
        else:
            target = kpi_tracker.objects.raw("""SELECT id, `name`, target, tier1, tier2, tier3, bonustarget, bonustier1,
                                                   bonustier2, bonustier3, bonustier4 FROM ced_kpi_tracker
                                                   WHERE contract_id = '%s' and tool = "PACO Rich" ORDER BY sortorder""" % (contract[0].id))
            tgt["home"] = list()
            tgt["office"] = list()
            for i in range(0, 3):
                tgt["home"].append(dict())
                tgt["home"][i]["tgt"] = target[i].target
                tgt["home"][i]["tier1"] = target[i].tier1
                tgt["home"][i]["tier2"] = target[i].tier2
                tgt["home"][i]["tier3"] = target[i].tier3
            for i in range(0, 2):
                tgt["office"].append(dict())
                if contract[0].id == 1:
                    try:
                        tgt["office"][0]["tgt"] = target[i+3].target
                        tgt["office"][0]["tier1"] = target[i+3].tier1
                        tgt["office"][0]["tier2"] = target[i+3].tier2
                        tgt["office"][0]["tier3"] = target[i+3].tier3
                    except IndexError:
                        None
                elif contract[0].id == 4:
                        tgt["office"][i]["tgt"] = target[i+3].target
                        tgt["office"][i]["tier1"] = target[i+3].tier1
                        tgt["office"][i]["tier2"] = target[i+3].tier2
                        tgt["office"][i]["tier3"] = target[i+3].tier3
        if (form["month"] == "" and (form["date_from"] == "" or form["date_to"] == "")):
            errors["empty"] = "Inserisci una Data"
        elif date_to < date_from:
            errors["date"] = "Data Futura o non Congruente"
        elif (form["month"] != "") and (int(str(form["Year"]) + month[0].number) > int(str(dt.datetime.now().year) + str(dt.datetime.now().month).zfill(2))):
            errors["date"] = "Data Futura o non Congruente"
        elif errors == {}:
            for i in range(0, len(table)):
                years = []
                data = []
                if form["month"] != "":
                    mcount = 1
                    fweek = int(dt.date(form["Year"], int(month[0].number), 1).strftime("%W"))
                    try:
                        lweek = int((dt.date(form["Year"], int(month[0].number) + 1, 1)-dt.timedelta(1)).strftime("%W"))
                    except ValueError:
                        lweek = 1
                    datef = "01" + "-" + str(int(month[0].number)) + "-" + str(form["Year"])
                    lastday = month_tr.objects.raw("SELECT name, days FROM ced_month_tr WHERE number = '%s'" % (month[0].number))
                    datel = str(lastday[0].days) + "-" + str(int(month[0].number)) + "-" + str(form["Year"])
                    if form["group"] == "ENNOVA":
                        df = pd.read_sql("SELECT * FROM %s WHERE substring(Data,4, 7) = '%s'"
                                        % (table[i], (month[0].number+"-"+str(form["Year"])[2:])), connection)
                    else:
                        df = pd.read_sql("SELECT * FROM %s WHERE Keys_BACINO_id = '%s' AND substring(Data,4, 7) = '%s'"
                                        % (table[i], group[0].ID, (month[0].number+"-"+str(form["Year"])[2:])), connection)
                    data.append(month[0].number + "-" + str(form["Year"][2:]))
                    years.append(str(form["Year"]))
                    columns = df.columns.tolist()
                else:
                    date_from = dt.datetime.strptime(form["date_from"], "%Y-%m-%d")
                    date_to = dt.datetime.strptime(form["date_to"], "%Y-%m-%d")
                    if int(form["date_from"][:4]) == int(form["date_to"][:4]):
                        month = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE number >= '%s' AND number <= '%s'ORDER BY number"
                                                    % (form["date_from"][5:7], form["date_to"][5:7]))
                        for mesi in month:
                            data.append(month[0].number + "-" + form["date_from"][2:4])
                        years.append(form["date_from"][:4])
                    else:
                        month1 = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE number >= '%s' AND number <= '12' ORDER BY number" % (form["date_from"][5:7]))
                        month2 = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE number >= '01' AND number <= '%s' ORDER BY number" % (form["date_to"][5:7]))
                        for mesi in month1:
                            data.append(mesi.number + "-" + form["date_from"][2:4])
                        for mesi in month2:
                            data.append(mesi.number + "-" + form["date_to"][2:4])
                        years.append(form["date_from"][:4])
                        years.append(form["date_to"][:4])
                    mcount = len(data)
                    datef = str(form["date_from"])[8:10] + "-" + str(form["date_from"][5:7]) + "-" + str(form["date_from"][2:4])
                    datel = str(form["date_to"])[8:10] + "-" + str(form["date_to"][5:7]) + "-" + str(form["date_to"][2:4])
                    if form["group"] == "ENNOVA":
                        df = pd.read_sql("SELECT * FROM " + table[i] + " WHERE str_to_date(Data,'%d-%m-%Y') >= '" + datef + "' AND str_to_date(Data,'%d-%m-%Y') <= '"
                                         + datel + "'", connection)
                    else:
                        df = pd.read_sql("SELECT * FROM " + table[i] + " WHERE Keys_BACINO_id = '" + group[0].ID +
                                         "' AND str_to_date(Data,'%d-%m-%Y') >= '" + datef + "' AND str_to_date(Data,'%d-%m-%Y') <= '"
                                         + datel + "'", connection)
                    columns = df.columns.tolist()
                    fweek = int(date_from.strftime("%W"))
                    lweek = int(date_to.strftime("%W"))
                if df.empty:
                    errors["empty"] = "Non ci sono dati nel periodo selezionato"
                else:
                    info = "NB. Selezionando un mese o un range di date che non parta dal lunedì e non finisca Domenica, la prima e l'ultima settimana potrebbero avere giorni mancanti."
                    richmonth[table[i]].append(dfo.richmonth(df, table[i], data, group[0].ID, columns, contract).copy())
                    richweek[table[i]].append(dfo.richweek(df, table[i], fweek, lweek, years,  group[0].ID, columns, contract).copy())
                    if richtable == "1":
                        filename = "Dettaglio Recalls 187 191 " + form["group"] + ".xlsx"
                        response = HttpResponse(content_type='application/vnd.ms-excel')
                        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
                        for k in range(0, len(table)):
                            querydown = qu.paco_download(table[k], data, groups[0].ID)
                            dfdown = pd.read_sql(querydown, connection)
                            dfdown.sort_values(by=["Keys_SERVIZIO_id"], ascending=False, inplace=True)
                            dfdown.set_index
                            dfdown.rename(columns={"Keys_SERVIZIO_id": "SERVICE", "Keys_BACINO_id": "BACINO"}, inplace=True)
                            dfdown.drop(["ID", "BACINO", "SERVICE", "Keys_ASC_id"], axis=1, inplace=True)
                            dflistdown.append(dfdown)
        tablerecall = fe.recalls_xls_template(table, dflistdown, datef, datel, htmlth, richmonth, richweek, tgt, contract)
        tablerecall.save(response)
        return response
    except Exception as error:
        print(type(error), error)
    context = {0: {'form': form},
               1: {'groups': groups},
               2: {'months': months},
               3: {'errors': errors},
               4: {'htmlth': htmlth},
               5: {'richmonth': richmonth},
               7: {'richweek': richweek},
               9: {'target': tgt},
               10: {'info': info},
               11: {'contract': contract}}

    return render(request, 'CED/richiamate.html', context)


@login_required
def kpi_c87(request):
    mdict = dict()
    target = dict()
    updatekpi = ""
    wb = Workbook()


    typeupd = ["FEH", "FEO", "BO"]

    dbcolumn = tz.columnkpo
    form = dict()
    errors = dict()
    groups = c87_group.objects.raw("SELECT * FROM ced_c87_group WHERE substring(BACINO, 1, 9) = 'ENNOVA_CA' OR BACINO = 'ENNOVA' ORDER BY BACINO")
    months = month_tr.objects.raw("SELECT name, number FROM ced_month_tr ORDER BY number")
    order = order_tracker.objects.raw("SELECT * FROM ced_order_tracker WHERE pafdesc = 'C87'")
    lengroup = len(list(groups))
    dfdict = dict()
    tables = ["ced_c87_ivr_187", "ced_c87_ivr_191", "ced_opera_sintesi_volumi_bo_home",
              "ced_opera_sintesi_volumi_bo_office", "ced_opera_sintesi_volumi_fe_home",
              "ced_opera_sintesi_volumi_fe_office", "ced_paco_richiamate_fe_187",
              "ced_paco_richiamate_fe_191"]

    try:
        form["year"] = request.POST["year"]
        form["group"] = request.POST["group"]
        form["tablekpi"] = request.POST["tablekpi"]
        form["month"] = request.POST["month"]

        try:
            form["sv_cc"] = request.POST["sv_cc"]
        except:
            form["sv_cc"] = 0
        try:
            form["all_group"] = request.POST["all_group"]
        except:
            form["all_group"] = 0

        month = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE name = '%s'" % form["month"])
        mdict[month[0].name] = month[0].number
        group = c87_group.objects.raw("SELECT ID, BACINO FROM ced_c87_group WHERE BACINO = '%s'" % form["group"])
        sqldate = dt.datetime.strptime(form["year"] + month[0].number + "01", "%Y%m%d")
        contract = contract_tracker.objects.raw("""SELECT * FROM ced_contract_tracker WHERE ced_contract_tracker.order_id = '%s'
                                                   AND fromdate <= '%s' AND todate >= '%s' ORDER BY id""" % (order[0].id, sqldate, sqldate))
        target = kpi_tracker.objects.raw("""SELECT id, `name`, target, tier1, tier2, tier3, bonustarget, bonustier1,
                                                   bonustier2, bonustier3, bonustier4 FROM ced_kpi_tracker
                                                   WHERE contract_id = '%s' ORDER BY sortorder""" % (contract[0].id))
        date = month[0].number + form["year"][2:]
        dateivr = []
        dateivr.append(month[0].number + "/" + form["year"])
        today_year = str(dt.datetime.now().year)
        today_month = str(dt.datetime.now().month)
        if int(date[2:]+date[:2]) > int(today_year[2:]+today_month.zfill(2)):
            errors["date"] = "Data futura o non congruente"
        elif len(list(contract)) == 0:
            errors["empty"] = "Contratto non presente per la data indicata, carica il contratto dalla pagina di amministrazione"
        elif len(list(contract)) > 1:
            errors["date"] = "Presenti più contratti per la data indicata, verifica i contratti dalla pagina di amministrazione"
        elif len(list(target)) == 0:
            errors["empty"] = "Target non presenti per il contratto in vigore, verifica target e contratti dalla pagina di amministrazione"
        if errors == {}:
            if (form["tablekpi"] == "1" or form["tablekpi"] == "2") and form["all_group"] == "1":
                group = groups
            for grcount in range(0, len(list(group))):
                if grcount > 0:
                    ivrsmonth = dict()
                    updatekpi = dict()
                    updatekpi["FEH"] = list()
                    updatekpi["FEO"] = list()
                    updatekpi["BO"] = list()

                for i in range(0, len(tables)):
                    if i < 2:
                        field = tz.ivrdb
                    elif i > 5:
                        field = tz.pacodb
                    elif i > 1:
                        field = tz.operadb
                    query = qu.dlselect_generic(tables[i], field, date, group, grcount)
                    if field == tz.operadb:
                        dfdict[tables[i]] = pd.read_sql(query, connection, index_col="Keys_SERVIZIO_id")
                    else:
                        dfdict[tables[i]] = pd.read_sql(query, connection)
                    if form["all_group"] != "1" and dfdict[tables[i]].empty:
                        errors["empty"] = "Non ci sono dati nel periodo selezionato o i dati non sono completi"

                #---
                lastday = c87_KPI_KPO_obtained.objects.raw("SELECT ID, max(day) FROM ced_c87_kpi_kpo_obtained WHERE MonthYear = '%s' AND substring(id,1,3) = '%s'" % (date, group[grcount].ID))
                if contract[0].id == 4:
                    updatekpi = vo.c872018_kpi(dfdict, tables, mdict, group, grcount, target, typeupd, lastday, date, dateivr, contract)
                elif contract[0].id == 1:
                    updatekpi = vo.c872017_kpi(dfdict, tables, mdict, group, grcount, target, typeupd, lastday, date, dateivr, contract)
                #---

                if form["tablekpi"] == "1" and errors == {}:
                    mails = 0
                    if updatekpi != {}:
                        htmltemplate = mt.header("KPI C87 ", month[0].name) + mt.kpo_template(updatekpi, typeupd, form, group, grcount) + mt.footer()
                        sender = tools_credentials.objects.raw("SELECT * FROM ced_tools_credentials WHERE description = 'Mail'")
                        message = MIMEMultipart("alternative", None, [MIMEText(htmltemplate, 'html')])
                        message['Subject'] = "KPI C87 " + group[grcount].BACINO + " " + date[:2] + "/" + date[2:]
                        message['From'] = "CED Cagliari Mailer"
                        message['To'] = ", ".join(lp.dl_kpi)
                        receiver = lp.dl_kpi
                        if form["sv_cc"] == "1":
                            message['Cc'] = ", ".join(lp.c87_supervisor)
                            receiver = lp.dl_kpi+lp.c87_supervisor
                        server = smtplib.SMTP_SSL(lp.smtpsrv, port=465)
                        server.login(sender[0].user, sender[0].password)
                        server.sendmail(sender[0].user, receiver, message.as_string())
                        server.quit()
                        mails += 1
                if form["tablekpi"] == "2" and errors == {}:
                    if updatekpi != {}:
                        wb = fe.c87_kpi_xls_template(wb, form, updatekpi, group, grcount)
            if form["tablekpi"] == "2" and errors == {}:
                filename = "Dettaglio KPI 187 191.xlsx"
                try:
                    key = '%016x' % random.getrandbits(64)
                    os.makedirs("./CED/downloads/" + key)
                    wb.save("./CED/downloads/" + key + "/" + filename)
                    request.session["filename"] = filename
                    request.session["dir"] = "./CED/downloads/" + key + "/"
                    return render(request, 'CED/download.html')
                except Exception as error:
                    print(type(error), error)
    except:
        form["tablekpi"] = 0
    if form["tablekpi"] == "1" and errors == {}:
        errors["empty"] = str(mails) + " Mail Inviate"
    context = {0: {'groups': groups},
               4: {'months': months},
               2: {'errors': errors},
               3: {'updatekpi': updatekpi},
               7: {'form': form}, 8: {'lengroup': lengroup}}

    return render(request, 'CED/kpi_c87.html', context)


@login_required
def kpi_ibia(request):

    form = dict()
    form["table"] = ""
    groups = dict()
    ibia = dict()
    dbtable = ["ced_c87_ibia_residential", "ced_c87_ibia_business"]
    tables = ["Home", "Bus"]
    ibiatz = tz.ibia_kpi_targets

    ASCS = c87_ASC.objects.raw("SELECT ID, ced_c87_asc.ASC FROM ced_c87_asc WHERE ID <> '00'")
    rowdb = c87_ibia_rows.objects.raw("""SELECT ID, ced_c87_ibia_rows.ROW FROM ced_c87_ibia_rows
                                         WHERE ID < 20 ORDER BY ID""")

    try:
        form["table"] = request.POST["table"]
        form["group"] = request.POST["group"]
    except:
        None
    try:
        if request.POST["filter"] == "1":
            form["asc"] = request.POST["asc"]
            if form["asc"] != "ENNOVA":
                groups = c87_group.objects.raw("""SELECT ced_c87_group.*, ced_c87_asc.* FROM ced_c87_group
                                                  JOIN ced_c87_asc ON ced_c87_group.Keys_ASC_id = ced_c87_asc.ID
                                                  WHERE ced_c87_asc.ASC = '%s' ORDER BY ced_c87_group.BACINO""" % form["asc"])
            else:
                groups = c87_group.objects.raw("SELECT * FROM ced_c87_group ORDER BY BACINO")
    except:
        None
    if form["table"] == "0":
        for i in range(0, len(dbtable)):
            #today = dt.datetime.now().strftime("%d/%m/%Y")
            today = "28/09/2018"
            group = c87_group.objects.raw("SELECT ID, BACINO FROM ced_c87_group WHERE BACINO = '%s'" % form["group"])
            if form["group"] != "ENNOVA":
                query = ("""SELECT * FROM %s JOIN ced_c87_ibia_rows ON ced_c87_ibia_rows.ID = DIAGNOSI_FE_MACRODES_id
                            WHERE Date = '%s' AND Keys_BACINO_id = '%s' AND DIAGNOSI_FE_MACRODES_id < 20"""
                            % (dbtable[i], today, group[0].ID))
            else:
                query = ("""SELECT * FROM %s JOIN ced_c87_ibia_rows ON ced_c87_ibia_rows.ID = DIAGNOSI_FE_MACRODES_id
                            WHERE Date = '%s' AND DIAGNOSI_FE_MACRODES_id <> '10' AND DIAGNOSI_FE_MACRODES_id < 20""" % (dbtable[i], today))
            df = pd.read_sql(query, connection)
            ibia[tables[i]] = dfo.ibia(df, group, today, rowdb, ibiatz, tables[i])
    context = {0: {'groups': groups}, 1: {'ASCS': ASCS}, 2: {'ibia': ibia},
               7: {'form': form}}
    return render(request, 'CED/kpi_ibia.html', context)


@login_required
def sos_easy_paf(request):
    form = dict()
    mdict = dict()
    errors = dict()
    paf = list()
    month = dict()
    svdata = dict()
    weektotal = list()
    total = 0
    tkts = 0
    months = month_tr.objects.raw("SELECT name, number FROM ced_month_tr ORDER BY number")
    order = order_tracker.objects.raw("SELECT * FROM ced_order_tracker WHERE pafdesc = 'ATP'")
    days = days_tr.objects.raw("SELECT * FROM ced_days_tr ORDER BY isonumber")

    try:
        form["year"] = request.POST["year"]
        form["month"] = request.POST["month"]
        form["tablepaf"] = request.POST["tablepaf"]
        month = month_tr.objects.raw("SELECT name, number, days FROM ced_month_tr WHERE name = '%s'" % form["month"])
        sqldate = dt.datetime.strptime(form["year"] + month[0].number + "01", "%Y%m%d")
        contract = contract_tracker.objects.raw("""SELECT * FROM ced_contract_tracker WHERE ced_contract_tracker.order_id = '%s'
                                           AND fromdate <= '%s' AND todate >= '%s' ORDER BY id""" % (order[0].id, sqldate, sqldate))
        tkts = tkt_money_value.objects.raw("""SELECT * FROM ced_tkt_money_value WHERE ced_tkt_money_value.contract_id = '%s'
                                              ORDER BY sortorder""" % (contract[0].id))
        svdata = supervisor_tracker.objects.raw("""SELECT * FROM ced_supervisor_tracker WHERE order_id = '%s'
                                                AND fromdate <= '%s' AND todate >= '%s' ORDER BY id""" % (order[0].id, sqldate, sqldate))
        sitedata = sites_tracker.objects.raw("""SELECT * FROM ced_sites_tracker WHERE id = '%s'""" % (order[0].site_id))
        festdata = festivity_tr.objects.raw("""SELECT * FROM ced_festivity_tr WHERE monthname_id = '%s'""" % (month[0].name))

        discountdata = pd.read_sql("""SELECT * FROM ced_tkt_money_value INNER JOIN (ced_discount_tracker INNER JOIN
                                      ced_tkt_discount_association ON ced_discount_tracker.id = ced_tkt_discount_association.discount_id)
                                      ON ced_tkt_money_value.id = ced_tkt_discount_association.tkt_id
                                      WHERE ced_tkt_money_value.contract_id = %s and todate >= '%s' and fromdate <= '%s'
                                      order by ced_tkt_money_value.sortorder, calc_order""" % (contract[0].id, sqldate, sqldate), connection)
        date = month[0].number + form["year"][2:]
        today_year = str(dt.datetime.now().year)
        today_month = str(dt.datetime.now().month)
        # today = int(form["year"] + month[0].number + str(dt.datetime.now().day).zfill(2))
        if int(date[2:]+date[:2]) > int(today_year[2:]+today_month.zfill(2)):
            errors["date"] = "Data futura o non congruente"
        if errors == {}:
            query = qu.dlselect_generic("ced_sos_easy_ibia", tz.sosibiadb, form["year"] + month[0].number)
            df = pd.read_sql(query, connection)
            if df.empty:
                errors["empty"] = "Non ci sono dati nel periodo selezionato"
            else:
                query = ("""SELECT DISTINCT DATA_TK, ISO_WEEK_DAY FROM ced_sos_easy_ibia
                         WHERE substring(DATA_TK, 1, 6) = '%s'""" % (form["year"] + month[0].number))
                dfdays = pd.read_sql(query, connection)
                paf = dfo.sos_easy_consutivo(df, dfdays, form["year"], month, days, tkts, discountdata, festdata)
            discdf = discountdata[(discountdata["pafdesc"] == tkts[0].pafdesc) & (discountdata["type_id"] == 2)].reset_index(drop=True)
            if form["tablepaf"] == "1":
                wb = fe.atpay_easy_paf(paf, str(month[0].days), svdata[0].surname, discdf["disc_value"][0], sitedata[0].sitename)
                filename = "PAF ATPay.xlsx"
                print(1)
                try:
                    key = '%016x' % random.getrandbits(64)
                    os.makedirs("./CED/downloads/" + key)
                    wb.save("./CED/downloads/" + key + "/" + filename)
                    request.session["filename"] = filename
                    request.session["dir"] = "./CED/downloads/" + key + "/"
                    return render(request, 'CED/download.html')
                except Exception as error:
                    print(type(error), error)
    except Exception as error:
        print(type(error), error)
    context = {2: {'errors': errors}, 3: {'month': month}, 4: {'months': months},
               5: {'paf': paf}, 6: {'weektotal': weektotal}, 7: {'form': form}, 8: {'days': days}, 9: {'svdata': svdata}}
    return render(request, 'CED/sos_easy_paf.html', context)


@login_required
def sos_easy_sms(request):
    form = dict()
    smsmonth = dict()
    errors = dict()
    smsweek = dict()
    df = []
    columns = []
    dates = []
    date_from = 0
    date_to = 0
    smstable = []
    info = ""
    htmlth = ["Periodo", "Fatturati Online", "Fake", "%Fake", "Non Presenti", "Totale", "KO", "%KO su Risposta", "Totale (singolo TT)", "KO (singolo TT)", "%"]
    table = ["ced_sos_easy_sms_fake", "ced_sos_easy_sms_non3xx", "ced_sos_easy_sms_risposte"]

    for i in range(0, len(table)):
        smsmonth[table[i]] = []
        smsweek[table[i]] = []

    info = "Attenzione, la prima e l'ultima settimana potrebbero avere giorni mancanti. Per avere una visione più precisa selezionare un range di date."

    months = month_tr.objects.raw("SELECT name, number FROM ced_month_tr ORDER BY number")
    try:
        form["year"] = int(request.POST["year"])
        smstable = request.POST["btn"]
        try:
            form["month"] = request.POST["month"]
            form["date_from"] = request.POST["date_from"]
            form["date_to"] = request.POST["date_to"]
            date_from = dt.datetime.strptime(form["date_from"], "%Y-%m-%d")
            date_to = dt.datetime.strptime(form["date_to"], "%Y-%m-%d")
            month = month_tr.objects.raw("SELECT name, number FROM ced_month_tr WHERE number >= '%s' AND number <= '%s'ORDER BY number"
                                        % (form["date_from"][5:7], form["date_to"][5:7]))
        except:
            form["month"] = request.POST["month"]
            month = month_tr.objects.raw("SELECT name, number, days FROM ced_month_tr WHERE name = '%s'" % form["month"])
        if (form["month"] == "" and (form["date_from"] == "" or form["date_to"] == "")):
            errors["empty"] = "Inserisci una Data"
        elif date_to < date_from:
            errors["date"] = "Data Futura o non Congruente"
        elif int(str(form["year"]) + month[0].number) > int(str(dt.datetime.now().year) + str(dt.datetime.now().month).zfill(2)):
            errors["date"] = "Data Futura o non Congruente"
        else:
            if form["month"] != "":
                fweek = int(dt.date(form["year"], int(month[0].number), 1).strftime("%W"))
                lweek = int((dt.date(form["year"], int(month[0].number) + 1, 1)-dt.timedelta(1)).strftime("%W"))
                datef = "01" + "-" + str(int(month[0].number)) + "-" + str(form["year"])
                lastday = month_tr.objects.raw("SELECT name, days FROM ced_month_tr WHERE number = '%s'" % (month[0].number))
                datel = str(lastday[0].days) + "-" + str(int(month[0].number)) + "-" + str(form["year"])
                dates.append(month[0].number + str(form["year"])[2:])
                for i in range(0, len(table)):
                    if i == 0:
                        field = tz.fakedb
                    elif i == 1:
                        field = tz.non3xxdb
                    elif i == 2:
                        field = tz.rispostedb
                    query = qu.dlselect_generic(table[i], field, dates)
                    df.append(pd.read_sql(query, connection))
                    columns.append(df[i].columns.tolist())
                querypaf = ("SELECT *, date_format(str_to_date(DATA_TK, '%Y%m%d'), '%v') as WEEK FROM ced_sos_easy_ibia WHERE substring(DATA_TK, 1, 6) = '" + (str(form["year"]) + str(month[0].number)) + "' AND "
                            "TIPO_INTERVENTO = 'ASSISTENZA ONLINE' AND tipo_lavorazione NOT LIKE '%FREE%' AND fatturato = 'si'")
                dfibia = pd.read_sql(querypaf, connection)
            else:
                fweek = int(date_from.strftime("%W"))
                lweek = int(date_to.strftime("%W"))
                dates.append(str(form["date_from"])[8:10] + month[0].number + str(form["year"])[2:])
                dates.append(str(form["date_to"])[8:10] + month[0].number + str(form["year"])[2:])
                datef = str(form["date_from"])[8:10] + "-" + month[0].number + "-" + str(form["year"])
                datel = str(form["date_to"])[8:10] + "-" + month[0].number + "-" + str(form["year"])
                for i in range(0, len(table)):
                    if i == 0:
                        field = tz.fakedbw
                    elif i == 1:
                        field = tz.non3xxdbw
                    elif i == 2:
                        field = tz.rispostedbw
                    query = qu.dlselect_generic(table[i], field, dates)
                    df.append(pd.read_sql(query, connection))
                    columns.append(df[i].columns.tolist())
                querypaf = ("SELECT *, date_format(str_to_date(DATA_TK, '%Y%m%d'), '%v') as WEEK FROM ced_sos_easy_ibia WHERE substring(DATA_TK, 1, 8) >= '" + (str(form["year"]) + str(month[0].number) + str(form["date_from"])[8:10]) + "' AND "
                            "substring(DATA_TK, 1, 8) <= '" + (str(form["year"]) + str(form["date_to"])[5:7] + str(form["date_to"])[8:10]) + "' AND TIPO_INTERVENTO = 'ASSISTENZA ONLINE' AND tipo_lavorazione NOT LIKE '%FREE%' AND fatturato = 'si'")
                dfibia = pd.read_sql(querypaf, connection)
            if df[0].empty:
                errors["empty"] = "Non ci sono dati nel periodo selezionato"
            else:
                if smstable == "0":
                    info = "NB. Selezionando un mese o un range di date che non parta dal lunedì e non finisca Domenica, la prima e l'ultima settimana potrebbero avere giorni mancanti."
                    smsmonth[table[i]] = dfo.smscalcmonth(df, dfibia, month, columns)
                    smsweek[table[i]] = dfo.smscalcweek(df, dfibia, fweek, lweek, columns)
                elif smstable == "1":
                    smsmonth[table[i]] = dfo.smscalcmonth(df, dfibia, month, columns)
                    smsweek[table[i]] = dfo.smscalcweek(df, dfibia, fweek, lweek, columns)
                    if form["month"] != "":
                        filename = "Sondaggi_Atpay_" + form["month"] + ".xlsx"
                    else:
                        filename = "Sondaggi_Atpay_dal_" + form["date_from"] + "_al_" + form["date_to"] + ".xlsx"
                    try:
                        filesms = fe.Atpay_easy_xls_template(htmlth, table, smsmonth, smsweek, datef, datel)
                        response = HttpResponse(content_type='application/vnd.ms-excel')
                        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
                        filesms.save(response)
                    except Exception as error:
                        print(type(error), error)
                    return response
    except:
        None
    context = {0: {'form': form},
               1: {'months': months},
               2: {'errors': errors},
               3: {'smsmonth': smsmonth},
               4: {'smsweek': smsweek},
               5: {'htmlth': htmlth},
               6: {'info': info}}

    return render(request, 'CED/sos_easy_sms.html', context)


@login_required
def under_constr(request):
    return render(request, 'CED/under_constr.html')
